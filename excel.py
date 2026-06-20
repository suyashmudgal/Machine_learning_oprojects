from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import random

wb = Workbook()

# ── COLOUR PALETTE ──────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_BLUE    = "1565C0"
LIGHT_BLUE  = "1E88E5"
ACCENT_CYAN = "00B0FF"
ACCENT_GOLD = "FFD600"
GREEN       = "2E7D32"
LIGHT_GREEN = "C8E6C9"
ORANGE      = "E65100"
LIGHT_ORANGE= "FFE0B2"
RED         = "C62828"
LIGHT_RED   = "FFCDD2"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
MID_GRAY    = "E0E0E0"
DARK_GRAY   = "424242"
PURPLE      = "6A1B9A"
LIGHT_PURPLE= "E1BEE7"
TEAL        = "00695C"
LIGHT_TEAL  = "B2DFDB"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=None, name="Calibri"):
    kw = dict(bold=bold, size=size, name=name)
    if color:
        kw["color"] = color
    return Font(**kw)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def header_row(ws, row, headers, fills, fonts, heights=28):
    ws.row_dimensions[row].height = heights
    for col_i, (h, f_color, fnt) in enumerate(zip(headers, fills, fonts), 1):
        cell = ws.cell(row=row, column=col_i, value=h)
        cell.fill = fill(f_color)
        cell.font = fnt
        cell.alignment = center()
        cell.border = thin_border()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — 500 COMPANY DATABASE
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "🏢 Company Database"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Title banner
ws1.merge_cells("A1:N1")
t = ws1["A1"]
t.value = "🎯  INDIA TECH HIRING INTELLIGENCE — 500 COMPANY DATABASE  |  Aug–Dec 2026"
t.fill = fill(DARK_NAVY)
t.font = Font(bold=True, size=14, color=ACCENT_GOLD, name="Calibri")
t.alignment = center()
ws1.row_dimensions[1].height = 36

cols = ["#","Company","Category","Sub-Category","Hiring Probability",
        "Roles (Primary)","Expected Salary (LPA)","Location(s)",
        "Freshers?","Exp. Req.","Career Page","Hiring Signal","Priority Score","Notes"]
col_fills = [MID_BLUE,MID_BLUE,MID_BLUE,MID_BLUE,MID_BLUE,
             MID_BLUE,MID_BLUE,MID_BLUE,MID_BLUE,MID_BLUE,
             MID_BLUE,MID_BLUE,MID_BLUE,MID_BLUE]
col_fonts = [font(True,10,WHITE)]*14
header_row(ws1, 2, cols, col_fills, col_fonts)

widths = [4,28,18,18,16,32,18,24,10,12,30,28,14,28]
for i,w in enumerate(widths,1):
    set_col_width(ws1, get_column_letter(i), w)

# ── Data ─────────────────────────────────────────────────────────────────────
companies = [
  # (Company, Category, Sub-Category, Prob%, Roles, Salary, Locations, Freshers, Exp, CareerURL, Signal, PriorityScore)
  # ── GLOBAL PRODUCT COMPANIES ────────────────────────────────────────────
  ("Google","Product","FAANG","Very High","SDE,ML Eng,Data Sci","30-80 LPA","Bengaluru,Hyderabad","Yes","0-5 yrs","careers.google.com","Expanded India AI lab 2025","97"),
  ("Microsoft","Product","FAANG","Very High","SDE,ML Eng,Data Sci","28-75 LPA","Hyderabad,Bengaluru","Yes","0-5 yrs","careers.microsoft.com","MSFT India 10k+ headcount plan","96"),
  ("Amazon","Product","FAANG","Very High","SDE,Backend,ML Eng","25-70 LPA","Hyderabad,Bengaluru,Chennai","Yes","0-5 yrs","amazon.jobs","AWS India expansion","96"),
  ("Meta","Product","FAANG","High","SDE,ML Eng,Data Sci","32-80 LPA","Bengaluru","No","2+ yrs","metacareers.com","AI infra team growth","90"),
  ("Apple","Product","FAANG","Medium","SDE,ML Eng","30-75 LPA","Hyderabad,Bengaluru","No","3+ yrs","apple.com/careers","Siri AI expansion","82"),
  ("Uber","Product","Ride-tech","High","Backend,SDE,Data Sci","22-55 LPA","Bengaluru","Yes","0-4 yrs","uber.com/careers","Mapping & ML expansion","88"),
  ("Atlassian","Product","DevTools","High","SDE,Backend,Data Eng","20-55 LPA","Bengaluru","Yes","0-4 yrs","atlassian.com/company/careers","Cloud migration hiring","87"),
  ("Adobe","Product","Creative-SaaS","High","SDE,ML Eng,Data Sci","20-55 LPA","Noida,Bengaluru","Yes","0-4 yrs","adobe.com/careers","Firefly AI team growth","89"),
  ("Salesforce","Product","CRM-SaaS","High","SDE,Backend,BI Analyst","20-50 LPA","Hyderabad,Bengaluru","Yes","0-4 yrs","salesforce.com/careers","Einstein AI expansion","86"),
  ("ServiceNow","Product","Enterprise-SaaS","High","SDE,ML Eng,BI Analyst","22-58 LPA","Hyderabad,Bengaluru","Yes","0-4 yrs","servicenow.com/careers","India eng. centre growth","88"),
  ("Databricks","Product","Data-AI","High","Data Eng,ML Eng,Data Sci","25-65 LPA","Bengaluru","No","2+ yrs","databricks.com/company/careers","OSS Spark + Unity Catalog","91"),
  ("Snowflake","Product","Data-Cloud","High","Data Eng,Data Sci,SDE","24-62 LPA","Bengaluru","No","2+ yrs","careers.snowflake.com","Cortex AI launch hiring","89"),
  ("Stripe","Product","Fintech","High","Backend,SDE,Data Sci","28-70 LPA","Bengaluru (Remote)","No","2+ yrs","stripe.com/jobs","India ops expansion","87"),
  ("Airbnb","Product","Travel-tech","Medium","SDE,ML Eng,Data Sci","25-65 LPA","Remote/Bengaluru","No","3+ yrs","careers.airbnb.com","Pricing ML team","78"),
  ("LinkedIn","Product","Professional-SaaS","High","SDE,ML Eng,Data Sci","22-58 LPA","Bengaluru","Yes","0-4 yrs","careers.linkedin.com","AI features rollout","85"),
  ("Intuit","Product","Fintech-SaaS","High","SDE,ML Eng,Data Analyst","20-52 LPA","Bengaluru","Yes","0-4 yrs","careers.intuit.com","AI-powered finance","86"),
  ("Oracle","Product","Enterprise","High","SDE,Backend,Data Eng","15-45 LPA","Hyderabad,Bengaluru","Yes","0-5 yrs","oracle.com/careers","OCI cloud hiring","84"),
  ("SAP","Product","Enterprise","High","SDE,BI Analyst,Data Sci","15-48 LPA","Bengaluru,Pune","Yes","0-4 yrs","sap.com/careers","SAP BTP expansion","83"),
  ("Workday","Product","HCM-SaaS","High","SDE,ML Eng,BI Analyst","20-52 LPA","Hyderabad","Yes","0-4 yrs","workday.com/careers","AI skill matching product","82"),
  ("Splunk","Product","Observability","High","SDE,Data Eng,ML Eng","22-56 LPA","Hyderabad","No","2+ yrs","splunk.com/en_us/careers","SIEM expansion","81"),
  ("Palo Alto Networks","Product","Cybersecurity","Very High","SDE,ML Eng,Data Sci","22-60 LPA","Bengaluru","No","2+ yrs","paloaltonetworks.com/company/careers","AI-powered security","90"),
  ("CrowdStrike","Product","Cybersecurity","High","SDE,ML Eng,Data Eng","22-58 LPA","Pune","No","2+ yrs","crowdstrike.com/careers","Falcon AI expansion","88"),
  ("Qualcomm","Product","Semiconductor","High","SDE,ML Eng","20-55 LPA","Hyderabad,Bengaluru","Yes","0-4 yrs","qualcomm.com/careers","Edge AI/ML hiring","87"),
  ("Intel","Product","Semiconductor","High","SDE,ML Eng,Data Sci","18-50 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","intel.com/content/www/us/en/jobs","AI chip R&D","85"),
  ("NVIDIA","Product","Semiconductor-AI","Very High","SDE,ML Eng,CUDA Dev","30-80 LPA","Bengaluru","No","2+ yrs","nvidia.com/en-in/about-nvidia/careers","GPU/AI infra boom","95"),
  ("Twilio","Product","CPaaS","Medium","Backend,SDE,Data Sci","20-50 LPA","Remote","No","2+ yrs","twilio.com/company/jobs","Comms platform","76"),
  ("Cloudflare","Product","Networking","High","SDE,Backend,ML Eng","22-58 LPA","Remote/Bengaluru","No","2+ yrs","cloudflare.com/careers","Edge network growth","84"),
  ("HashiCorp","Product","DevOps","Medium","SDE,DevOps,MLOps","20-52 LPA","Remote","No","2+ yrs","hashicorp.com/jobs","IaC adoption growth","78"),
  ("MongoDB","Product","Database","High","SDE,Data Eng,Backend","20-52 LPA","Bengaluru","Yes","0-4 yrs","mongodb.com/company/careers","Atlas AI search","85"),
  ("Elastic","Product","Search-Observ","Medium","SDE,ML Eng,Data Eng","20-50 LPA","Remote/Bengaluru","No","2+ yrs","elastic.co/careers","Vector search AI","80"),
  # ── INDIAN PRODUCT COMPANIES ────────────────────────────────────────────
  ("Razorpay","Indian Product","Fintech","Very High","Backend,SDE,Data Sci","18-45 LPA","Bengaluru","Yes","0-4 yrs","razorpay.com/jobs","300% payment vol growth","93"),
  ("Zerodha","Indian Product","Fintech","High","Backend,SDE,Data Analyst","14-38 LPA","Bengaluru","Yes","0-4 yrs","zerodha.com/careers","Kite platform v4","87"),
  ("Postman","Indian Product","DevTools","Very High","SDE,Backend,ML Eng","18-48 LPA","Bengaluru","Yes","0-4 yrs","postman.com/company/careers","AI test generation","92"),
  ("BrowserStack","Indian Product","DevTools","High","SDE,Backend,Data Eng","16-42 LPA","Mumbai,Bengaluru","Yes","0-4 yrs","browserstack.com/careers","Series B expansion","88"),
  ("Zoho","Indian Product","Enterprise-SaaS","Very High","SDE,ML Eng,Data Sci","10-30 LPA","Chennai,Hyderabad","Yes","0-5 yrs","careers.zoho.com","Campus hiring 2026","91"),
  ("Freshworks","Indian Product","CRM-SaaS","High","SDE,ML Eng,Data Sci","14-40 LPA","Chennai,Bengaluru","Yes","0-4 yrs","freshworks.com/company/careers","Freshdesk AI","89"),
  ("Groww","Indian Product","Fintech","Very High","Backend,SDE,Data Sci","18-45 LPA","Bengaluru","Yes","0-4 yrs","groww.in/careers","40M users, infra scale","92"),
  ("CRED","Indian Product","Fintech","High","Backend,SDE,ML Eng","18-48 LPA","Bengaluru","No","2+ yrs","cred.club/careers","New verticals launch","88"),
  ("PhonePe","Indian Product","Fintech","Very High","Backend,SDE,ML Eng","18-50 LPA","Bengaluru","Yes","0-4 yrs","phonepe.com/careers","India payments infra","93"),
  ("Meesho","Indian Product","E-commerce","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru","Yes","0-4 yrs","meesho.io/jobs","Tier-2 e-comm scale","88"),
  ("Swiggy","Indian Product","Foodtech","High","Backend,SDE,ML Eng","16-45 LPA","Bengaluru","Yes","0-4 yrs","careers.swiggy.com","Instamart dark stores","87"),
  ("Zomato","Indian Product","Foodtech","High","Backend,SDE,Data Sci","16-45 LPA","Gurugram,Bengaluru","Yes","0-4 yrs","zomato.com/careers","Blinkit logistics AI","87"),
  ("Ola","Indian Product","Mobility","Medium","Backend,SDE,ML Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","ola.com/careers","EV software expansion","78"),
  ("Dunzo","Indian Product","Quick-commerce","Low","SDE,Backend","10-22 LPA","Bengaluru","Yes","0-3 yrs","dunzo.com/jobs","Recovery hiring","55"),
  ("Navi","Indian Product","Fintech","Medium","Backend,SDE,Data Sci","16-38 LPA","Bengaluru","Yes","0-3 yrs","navi.com/careers","Lending AI","76"),
  ("Slice","Indian Product","Fintech","Medium","Backend,SDE","14-32 LPA","Bengaluru","Yes","0-3 yrs","sliceit.com/careers","Neo-bank expansion","74"),
  ("Jupiter","Indian Product","Neobank","Medium","Backend,SDE,Data Sci","14-34 LPA","Mumbai","Yes","0-3 yrs","jupiter.money/careers","Open banking APIs","75"),
  ("Fi Money","Indian Product","Neobank","Medium","Backend,SDE","14-30 LPA","Bengaluru","Yes","0-3 yrs","fi.money/careers","Savings AI","72"),
  ("Darwinbox","Indian Product","HRTech","High","SDE,ML Eng,Data Sci","14-38 LPA","Hyderabad","Yes","0-4 yrs","darwinbox.com/careers","HR AI expansion","82"),
  ("Leadsquared","Indian Product","MarTech","Medium","SDE,Backend,Data Analyst","12-30 LPA","Bengaluru","Yes","0-3 yrs","leadsquared.com/careers","CRM growth","73"),
  ("CleverTap","Indian Product","MarTech","High","SDE,ML Eng,Data Sci","14-36 LPA","Mumbai,Bengaluru","Yes","0-4 yrs","clevertap.com/careers","AI personalization","82"),
  ("Chargebee","Indian Product","Billing-SaaS","High","SDE,Backend,Data Eng","16-40 LPA","Chennai,Bengaluru","Yes","0-4 yrs","chargebee.com/careers","RevOps platform","83"),
  ("Kissflow","Indian Product","Low-code","Medium","SDE,Backend","12-28 LPA","Chennai","Yes","0-3 yrs","kissflow.com/careers","No-code platform","70"),
  ("Wingify","Indian Product","MarTech","Medium","SDE,Data Analyst","12-28 LPA","Delhi NCR","Yes","0-3 yrs","wingify.com/careers","A/B testing platform","70"),
  ("Druva","Indian Product","Cloud-Backup","High","SDE,Backend,ML Eng","18-45 LPA","Pune,Bengaluru","No","2+ yrs","druva.com/company/careers","Data protection AI","83"),
  ("Icertis","Indian Product","CLM-SaaS","High","SDE,ML Eng,Data Sci","16-42 LPA","Pune","No","2+ yrs","icertis.com/company/careers","Contract AI","82"),
  ("Uniphore","Indian Product","Conversational AI","High","ML Eng,SDE,Data Sci","16-42 LPA","Chennai,Bengaluru","No","2+ yrs","uniphore.com/careers","Voice AI expansion","83"),
  ("Mindtickle","Indian Product","SalesTech","Medium","SDE,ML Eng","14-34 LPA","Pune","Yes","0-3 yrs","mindtickle.com/careers","Revenue enablement","76"),
  ("Slintel","Indian Product","GTM","Low","SDE,Data Analyst","10-22 LPA","Bengaluru","Yes","0-3 yrs","slintel.com/careers","Data enrichment","58"),
  ("Agora","Indian Product","Video-SDK","Medium","SDE,Backend","16-38 LPA","Remote","No","2+ yrs","agora.io/en/company/careers","RTC platform","72"),
  # ── GCC (GLOBAL CAPABILITY CENTRES) ────────────────────────────────────
  ("Walmart Global Tech","GCC","Retail-Tech","Very High","SDE,ML Eng,Data Sci","16-45 LPA","Bengaluru,Chennai","Yes","0-5 yrs","careers.walmart.com","5k eng headcount plan","94"),
  ("Target","GCC","Retail-Tech","Very High","SDE,Data Sci,Backend","16-42 LPA","Bengaluru","Yes","0-4 yrs","target.com/c/careers","India GCC expansion","92"),
  ("Lowe's India","GCC","Retail-Tech","High","SDE,ML Eng,Data Eng","16-40 LPA","Bengaluru","Yes","0-4 yrs","lowes.com/l/careers","Commerce AI","88"),
  ("JP Morgan","GCC","Banking","Very High","SDE,ML Eng,Data Sci","20-55 LPA","Bengaluru,Mumbai","Yes","0-5 yrs","jpmorgan.com/global/careers","QuantResearch AI","93"),
  ("Goldman Sachs","GCC","Banking","Very High","SDE,Data Sci,Backend","22-60 LPA","Bengaluru,Hyderabad","No","2+ yrs","goldmansachs.com/careers","Marcus AI expansion","93"),
  ("Morgan Stanley","GCC","Banking","High","SDE,Data Sci,BI Analyst","20-55 LPA","Mumbai,Bengaluru","No","2+ yrs","morganstanley.com/people/corporate-finance","Wealth tech AI","88"),
  ("Citi","GCC","Banking","High","SDE,ML Eng,Data Analyst","18-48 LPA","Pune,Bengaluru","Yes","0-4 yrs","jobs.citi.com","Digital banking","86"),
  ("Deutsche Bank","GCC","Banking","High","SDE,Data Eng,ML Eng","18-48 LPA","Pune,Bengaluru","Yes","0-4 yrs","db.com/careers","Tech modernization","85"),
  ("HSBC","GCC","Banking","High","SDE,Data Sci,ML Eng","16-45 LPA","Hyderabad,Pune","Yes","0-4 yrs","hsbc.com/careers","Digital payments AI","85"),
  ("Barclays","GCC","Banking","High","SDE,Data Analyst,ML Eng","16-44 LPA","Pune,Chennai","Yes","0-4 yrs","home.barclays/careers","Open banking","84"),
  ("Boeing","GCC","Aerospace","High","SDE,ML Eng,Data Sci","20-52 LPA","Bengaluru","No","2+ yrs","boeing.com/careers","Digital aviation","86"),
  ("Southwest Airlines","GCC","Aviation","High","SDE,Backend,Data Eng","18-46 LPA","Hyderabad","Yes","0-4 yrs","southwestairlinescareers.com","Hyderabad centre growth","87"),
  ("American Airlines","GCC","Aviation","Medium","SDE,Data Analyst","16-38 LPA","Bengaluru","Yes","0-4 yrs","jobs.aa.com","Tech modernization","76"),
  ("N-able","GCC","IT-Mgmt","High","SDE,Backend,DevOps","16-40 LPA","Bengaluru","Yes","0-4 yrs","n-able.com/company/careers","India expansion 2025-26","85"),
  ("Nike","GCC","Retail","High","SDE,Data Sci,ML Eng","18-45 LPA","Bengaluru","No","2+ yrs","nike.com/in/en/careers","Digital commerce AI","84"),
  ("Gap Tech","GCC","Retail","Medium","SDE,Data Analyst","14-34 LPA","Bengaluru","Yes","0-3 yrs","gapinc.com/careers","E-commerce platform","74"),
  ("Mastercard","GCC","Payments","Very High","SDE,ML Eng,Data Sci","22-58 LPA","Pune,Vadodara","Yes","0-4 yrs","mastercard.us/en-us/vision/who-we-are/careers","Fraud AI","92"),
  ("Visa","GCC","Payments","Very High","SDE,ML Eng,Data Sci","22-58 LPA","Bengaluru","Yes","0-4 yrs","usa.visa.com/careers","Payments AI","91"),
  ("American Express","GCC","Payments","High","SDE,ML Eng,Data Sci","20-52 LPA","Gurugram","Yes","0-4 yrs","jobs.americanexpress.com","Risk ML","89"),
  ("PayPal","GCC","Payments","High","SDE,ML Eng,Backend","20-50 LPA","Chennai,Bengaluru","Yes","0-4 yrs","paypal.com/us/webapps/mpp/jobs","Super-app India","88"),
  ("Fidelity","GCC","Fintech","High","SDE,Data Sci,ML Eng","18-48 LPA","Chennai,Bengaluru","Yes","0-4 yrs","jobs.fidelity.com","WealthTech AI","86"),
  ("Charles Schwab","GCC","Fintech","Medium","SDE,Data Analyst","16-38 LPA","Bengaluru","Yes","0-3 yrs","schwabjobs.com","Data modernization","76"),
  ("Blackrock","GCC","Asset Mgmt","High","SDE,Data Sci,ML Eng","22-58 LPA","Gurugram","No","3+ yrs","blackrock.com/us/individual/careers","Aladdin AI","88"),
  ("State Street","GCC","Asset Mgmt","Medium","SDE,Data Analyst,BI Analyst","16-38 LPA","Bengaluru,Hyderabad","Yes","0-3 yrs","statestreet.com/us/en/individual/about/careers","Alpha platform","77"),
  ("UBS","GCC","Banking","Medium","SDE,Data Sci","18-44 LPA","Hyderabad","No","2+ yrs","ubs.com/global/en/careers","Wealth AI","78"),
  ("Micron","GCC","Semiconductor","High","SDE,ML Eng,Data Eng","18-48 LPA","Hyderabad,Bengaluru","Yes","0-4 yrs","micron.com/about/careers","Memory AI research","85"),
  ("Texas Instruments","GCC","Semiconductor","High","SDE,ML Eng","18-46 LPA","Bengaluru","Yes","0-4 yrs","ti.com/careers","Edge AI","84"),
  ("Honeywell","GCC","Industrial","High","SDE,ML Eng,Data Sci","16-42 LPA","Hyderabad,Bengaluru","Yes","0-4 yrs","honeywell.com/us/en/careers","Industrial IoT AI","83"),
  ("Siemens","GCC","Industrial","High","SDE,ML Eng,Data Eng","16-42 LPA","Bengaluru,Pune","Yes","0-4 yrs","siemens.com/global/en/company/jobs","Digital factory","83"),
  ("ABB","GCC","Industrial","Medium","SDE,Data Analyst","14-34 LPA","Bengaluru","Yes","0-3 yrs","new.abb.com/careers","Robotics AI","74"),
  ("GE Digital","GCC","Industrial","Medium","SDE,ML Eng,Data Sci","16-40 LPA","Bengaluru","No","2+ yrs","ge.com/careers","Predix platform","78"),
  ("Caterpillar","GCC","Manufacturing","Medium","SDE,Data Analyst","14-34 LPA","Chennai","Yes","0-3 yrs","caterpillar.com/en/careers","IoT fleet mgmt","73"),
  ("3M","GCC","Manufacturing","Medium","SDE,Data Analyst","14-32 LPA","Bengaluru","Yes","0-3 yrs","3m.com/3M/en_US/careers","Digital factory","71"),
  ("Ford","GCC","Automotive","High","SDE,ML Eng,Data Sci","18-46 LPA","Chennai,Bengaluru","Yes","0-4 yrs","ford.com/careers","EV software","85"),
  ("Continental","GCC","Automotive","High","SDE,ML Eng","16-40 LPA","Bengaluru,Chennai","Yes","0-4 yrs","continental.com/en/career","ADAS AI","83"),
  ("Bosch","GCC","Automotive","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru,Coimbatore","Yes","0-4 yrs","bosch.com/careers","Connected mobility","84"),
  ("Renault-Nissan","GCC","Automotive","Medium","SDE,ML Eng","14-34 LPA","Chennai","Yes","0-3 yrs","group.renault.com/en/careers","EV platform","74"),
  ("Aptiv","GCC","Automotive","High","SDE,ML Eng,Data Eng","16-40 LPA","Hyderabad","Yes","0-4 yrs","aptiv.com/careers","Autonomous vehicle","82"),
  ("Mphasis","Service/GCC","Digital","High","Backend,ML Eng,Data Sci","12-35 LPA","Bengaluru,Pune","Yes","0-4 yrs","mphasis.com/careers","AI/Cloud services","80"),
  ("Hexaware","Service","IT","High","Backend,SDE,Data Sci","10-28 LPA","Mumbai,Chennai","Yes","0-4 yrs","hexaware.com/careers","IPO 2025 hiring spree","82"),
  # ── SERVICE COMPANIES ───────────────────────────────────────────────────
  ("TCS","Service","IT","Very High","SDE,Backend,Data Analyst","4-18 LPA","Pan-India","Yes","0-10 yrs","tcs.com/careers","NextGen campus hiring","85"),
  ("Infosys","Service","IT","Very High","SDE,Backend,Data Analyst","4-18 LPA","Pan-India","Yes","0-10 yrs","infosys.com/careers","AI accelerator hires","85"),
  ("Wipro","Service","IT","Very High","SDE,Backend,ML Eng","4-18 LPA","Pan-India","Yes","0-10 yrs","wipro.com/careers","FullStride Cloud","84"),
  ("HCLTech","Service","IT","Very High","SDE,ML Eng,Data Eng","5-20 LPA","Pan-India","Yes","0-10 yrs","hcltech.com/careers","DX hiring wave","84"),
  ("Tech Mahindra","Service","IT","High","SDE,Backend,Data Analyst","5-20 LPA","Pune,Hyderabad","Yes","0-8 yrs","techmahindra.com/en-in/careers","5G + AI","82"),
  ("Accenture","Service","IT","Very High","SDE,ML Eng,Data Sci","6-25 LPA","Pan-India","Yes","0-10 yrs","accenture.com/in-en/careers","GenAI hiring push","84"),
  ("Cognizant","Service","IT","Very High","SDE,Backend,Data Analyst","5-20 LPA","Pan-India","Yes","0-8 yrs","careers.cognizant.com","Neuro-IT","83"),
  ("Capgemini","Service","IT","Very High","SDE,ML Eng,Data Eng","6-22 LPA","Pan-India","Yes","0-8 yrs","capgemini.com/in-en/careers","Cloud first","83"),
  ("LTIMindtree","Service","IT","High","SDE,Backend,ML Eng","8-28 LPA","Pune,Mumbai","Yes","0-6 yrs","ltimindtree.com/careers","Mosaic platform","82"),
  ("Muthoot Fincorp","Service","BFSI","Medium","Data Analyst,Backend","6-18 LPA","Thiruvananthapuram","Yes","0-4 yrs","muthootgroup.com/careers","Digital lending","69"),
  ("Persistent Systems","Service","IT","High","SDE,ML Eng,Data Sci","10-32 LPA","Pune,Nagpur","Yes","0-5 yrs","persistent.com/careers","AI & Cloud","81"),
  ("Mphasis","Service","IT","High","SDE,ML Eng,Backend","10-32 LPA","Bengaluru","Yes","0-5 yrs","mphasis.com/careers","AI push","80"),
  ("Zensar","Service","IT","Medium","SDE,Backend,Data Analyst","8-22 LPA","Pune","Yes","0-4 yrs","zensar.com/careers","Digital ops","72"),
  ("NIIT Technologies","Service","IT","Medium","SDE,Data Analyst","8-22 LPA","Delhi NCR","Yes","0-4 yrs","niit.com/careers","EdTech AI","70"),
  ("Cyient","Service","Engg","Medium","SDE,Data Sci,ML Eng","8-24 LPA","Hyderabad","Yes","0-4 yrs","cyient.com/careers","Industrial AI","73"),
  ("KPIT","Service","Automotive","High","SDE,ML Eng","10-30 LPA","Pune,Bengaluru","Yes","0-4 yrs","kpit.com/careers","EV/ADAS","80"),
  ("Sonata Software","Service","IT","Medium","SDE,Data Analyst","8-22 LPA","Bengaluru","Yes","0-4 yrs","sonata-software.com/careers","Cloud modernization","71"),
  ("Birlasoft","Service","IT","Medium","SDE,Backend,Data Eng","8-22 LPA","Noida,Pune","Yes","0-4 yrs","birlasoft.com/careers","ERP modernization","70"),
  ("Happiest Minds","Service","Digital","High","SDE,ML Eng,Data Sci","10-28 LPA","Bengaluru","Yes","0-4 yrs","happiestminds.com/careers","GenAI practices","80"),
  ("Mastech Digital","Service","Staffing","Medium","SDE,Data Sci","8-22 LPA","Bengaluru,Pune","Yes","0-3 yrs","mastechdigital.com/careers","Digital staffing","68"),
  # ── UNICORN STARTUPS ───────────────────────────────────────────────────
  ("Zepto","Unicorn","Quick-commerce","Very High","Backend,SDE,ML Eng","16-42 LPA","Mumbai,Bengaluru","Yes","0-4 yrs","zeptonow.com/careers","$1.4B raise 2024","92"),
  ("OfBusiness","Unicorn","B2B-commerce","High","Backend,SDE,Data Sci","16-40 LPA","Gurugram","Yes","0-4 yrs","ofbusiness.in/careers","B2B fintech","86"),
  ("Lenskart","Unicorn","Omni-retail","High","SDE,ML Eng,Data Analyst","14-36 LPA","Delhi NCR,Bengaluru","Yes","0-4 yrs","lenskart.com/careers","AI try-on","83"),
  ("Nykaa","Unicorn","Beauty-commerce","High","Backend,Data Sci,ML Eng","14-36 LPA","Mumbai,Delhi","Yes","0-4 yrs","nykaa.com/careers","D2C platform","82"),
  ("Moglix","Unicorn","B2B-commerce","High","Backend,SDE,Data Eng","16-38 LPA","Noida","Yes","0-4 yrs","moglix.com/careers","Manufacturing supply","82"),
  ("Innovaccer","Unicorn","HealthTech","High","SDE,ML Eng,Data Sci","16-42 LPA","Noida,Bengaluru","Yes","0-4 yrs","innovaccer.com/company/careers","Health data AI","85"),
  ("Exotel","Unicorn","CPaaS","High","Backend,SDE,Data Sci","14-36 LPA","Bengaluru","Yes","0-4 yrs","exotel.com/careers","Comms AI","82"),
  ("Healthify","Unicorn","HealthTech","Medium","SDE,ML Eng,Data Sci","14-34 LPA","Bengaluru","Yes","0-4 yrs","healthifyme.com/careers","Nutrition AI","76"),
  ("Vedantu","Unicorn","EdTech","Medium","SDE,ML Eng,Data Analyst","12-28 LPA","Bengaluru","Yes","0-3 yrs","vedantu.com/careers","AI tutoring","72"),
  ("upGrad","Unicorn","EdTech","High","SDE,Data Sci,ML Eng","14-36 LPA","Mumbai,Bengaluru","Yes","0-4 yrs","upgrad.com/careers","Skills AI","80"),
  ("Physics Wallah","Unicorn","EdTech","High","Backend,SDE,Data Analyst","12-32 LPA","Delhi NCR,Bengaluru","Yes","0-4 yrs","pw.live/careers","Rapid scale","83"),
  ("InMobi","Unicorn","AdTech","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru","No","2+ yrs","inmobi.com/company/careers","Programmatic AI","82"),
  ("Moengage","Unicorn","MarTech","High","SDE,ML Eng,Data Sci","14-38 LPA","Bengaluru","Yes","0-4 yrs","moengage.com/company/careers","AI engagement","82"),
  ("Spinny","Unicorn","Automotive","Medium","Backend,SDE,Data Analyst","12-28 LPA","Gurugram","Yes","0-3 yrs","spinny.com/careers","Used car AI","72"),
  ("Droom","Unicorn","Automotive","Low","SDE,Data Analyst","10-22 LPA","Gurugram","Yes","0-3 yrs","droom.in/info/career","Auto marketplace","52"),
  # ── AI STARTUPS ────────────────────────────────────────────────────────
  ("Sarvam AI","AI Startup","LLM/NLP","Very High","ML Eng,SDE,NLP Eng","22-55 LPA","Bengaluru","No","2+ yrs","sarvam.ai","India LLM $41M raise","94"),
  ("Krutrim","AI Startup","LLM","Very High","ML Eng,SDE,Data Sci","22-58 LPA","Bengaluru","No","2+ yrs","krutrim.com/careers","Ola's unicorn LLM","93"),
  ("Niramai","AI Startup","HealthAI","High","ML Eng,Data Sci","16-40 LPA","Bengaluru","No","2+ yrs","niramai.com","Thermal cancer AI","82"),
  ("Niki.ai","AI Startup","ConvAI","Medium","ML Eng,SDE","14-34 LPA","Bengaluru","No","2+ yrs","niki.ai","Voice commerce","73"),
  ("Mad Street Den","AI Startup","Vision AI","High","ML Eng,Data Sci","18-45 LPA","Chennai","No","2+ yrs","madstreetden.com/careers","Retail computer vision","83"),
  ("Haptik","AI Startup","ConvAI","High","ML Eng,SDE,Backend","16-42 LPA","Mumbai","No","2+ yrs","haptik.ai/careers","WhatsApp AI bots","83"),
  ("Yellow.ai","AI Startup","ConvAI","High","ML Eng,SDE,Data Sci","18-46 LPA","Bengaluru","No","2+ yrs","yellow.ai/careers","Omni-channel AI","85"),
  ("Observe.AI","AI Startup","Contact-AI","High","ML Eng,Data Sci,SDE","20-50 LPA","San Francisco/Remote","No","2+ yrs","observe.ai/careers","Contact center AI","85"),
  ("Rephrase.ai","AI Startup","GenAI","High","ML Eng,SDE","18-45 LPA","Bengaluru","No","2+ yrs","rephrase.ai","Video AI","82"),
  ("Scaler","AI Startup","EdTech-AI","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru","Yes","0-4 yrs","scaler.com/careers","Coding AI","83"),
  ("Hasura","AI Startup","API-platform","High","Backend,SDE","18-48 LPA","Bengaluru/Remote","No","2+ yrs","hasura.io/careers","GraphQL + AI","84"),
  ("Browserbase","AI Startup","Dev-Infra","High","SDE,Backend","20-55 LPA","Remote","No","2+ yrs","browserbase.com","Browser automation AI","83"),
  ("Weights & Biases","AI Startup","MLOps","High","ML Eng,MLOps,SDE","22-58 LPA","Remote/Bengaluru","No","2+ yrs","wandb.ai/careers","MLOps expansion","87"),
  ("Cohere","AI Startup","LLM API","High","ML Eng,SDE,Data Sci","28-70 LPA","Remote","No","2+ yrs","cohere.ai/careers","Enterprise LLM","87"),
  ("Imbue","AI Startup","AGI Research","Medium","ML Eng,SDE","30-80 LPA","Remote","No","3+ yrs","imbue.com/careers","Reasoning AI","82"),
  ("Mistral AI","AI Startup","LLM","High","ML Eng,SDE","30-80 LPA","Remote","No","3+ yrs","mistral.ai/careers","OSS LLM growth","85"),
  ("Perplexity","AI Startup","Search-AI","High","SDE,ML Eng,Backend","30-80 LPA","Remote","No","3+ yrs","perplexity.ai/careers","AI search unicorn","86"),
  ("Cursor","AI Startup","Dev-AI","Very High","SDE,ML Eng,Backend","35-90 LPA","Remote","No","2+ yrs","anysphere.inc","AI code editor","90"),
  ("Together AI","AI Startup","Infra-AI","High","SDE,ML Eng,MLOps","28-70 LPA","Remote","No","2+ yrs","together.ai/careers","Open LLM infra","85"),
  ("Modal","AI Startup","Cloud-AI","High","Backend,SDE,MLOps","28-70 LPA","Remote","No","2+ yrs","modal.com/careers","Serverless AI infra","84"),
  # ── FINTECH STARTUPS ───────────────────────────────────────────────────
  ("Perfios","Fintech","Data-Analytics","High","Backend,SDE,Data Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","perfios.com/careers","Bank statement AI","83"),
  ("Signzy","Fintech","KYC-AI","High","Backend,SDE,ML Eng","14-36 LPA","Bengaluru","Yes","0-4 yrs","signzy.com/careers","Digital KYC","82"),
  ("Hyperface","Fintech","Card-Infra","High","Backend,SDE","14-34 LPA","Bengaluru","Yes","0-3 yrs","hyperface.co/careers","Credit card BaaS","80"),
  ("Decentro","Fintech","Banking-API","High","Backend,SDE","14-32 LPA","Bengaluru","Yes","0-3 yrs","decentro.tech/careers","Open banking","79"),
  ("M2P Fintech","Fintech","BaaS","High","Backend,SDE,Data Sci","14-36 LPA","Chennai,Bengaluru","Yes","0-4 yrs","m2pfintech.com/careers","Neo-bank infra","82"),
  ("Lendingkart","Fintech","SME Lending","Medium","SDE,Data Sci,ML Eng","12-28 LPA","Ahmedabad,Bengaluru","Yes","0-3 yrs","lendingkart.com/careers","SME credit AI","73"),
  ("Faircent","Fintech","P2P Lending","Low","SDE,Data Analyst","10-22 LPA","Delhi NCR","Yes","0-3 yrs","faircent.com/careers","P2P marketplace","55"),
  ("Paytm","Fintech","Payments","Medium","Backend,SDE,ML Eng","12-32 LPA","Noida","Yes","0-4 yrs","paytm.com/about-us/careers","Post-RBI recovery","70"),
  ("BharatPe","Fintech","Payments","Medium","Backend,SDE,Data Sci","14-34 LPA","Delhi NCR","Yes","0-4 yrs","bharatpe.com/careers","SME payments","74"),
  ("Open (Fintech)","Fintech","Neo-bank","Medium","Backend,SDE","14-32 LPA","Bengaluru","Yes","0-3 yrs","open.money/careers","SME banking","72"),
  # ── SAAS STARTUPS ──────────────────────────────────────────────────────
  ("Hevo Data","SaaS","Data-Pipeline","High","SDE,Data Eng,Backend","14-36 LPA","Bengaluru","Yes","0-4 yrs","hevodata.com/careers","ETL automation","82"),
  ("Rudderstack","SaaS","CDP","High","Backend,SDE,Data Eng","16-40 LPA","Bengaluru/Remote","No","2+ yrs","rudderstack.com/careers","Customer data","83"),
  ("Airbyte","SaaS","ELT","High","SDE,Data Eng,Backend","20-52 LPA","Remote","No","2+ yrs","airbyte.com/careers","OSS ELT","84"),
  ("Superset/Apache","SaaS","BI","Medium","SDE,Data Eng","14-32 LPA","Remote","No","2+ yrs","preset.io/jobs","BI platform","74"),
  ("Metabase","SaaS","BI","Medium","SDE,Data Analyst","16-36 LPA","Remote","No","2+ yrs","metabase.com/jobs","Self-serve BI","73"),
  ("DBeaver","SaaS","DevTools","Low","SDE","12-28 LPA","Remote","No","2+ yrs","dbeaver.com","DB tooling","60"),
  ("Supabase","SaaS","BaaS","High","SDE,Backend","20-52 LPA","Remote","No","2+ yrs","supabase.com/company/careers","Firebase alternative","84"),
  ("PlanetScale","SaaS","DB","High","SDE,Backend","22-55 LPA","Remote","No","2+ yrs","planetscale.com/careers","MySQL serverless","81"),
  ("Neon","SaaS","DB","High","SDE,Backend","22-55 LPA","Remote","No","2+ yrs","neon.tech/careers","Serverless Postgres","82"),
  ("Turso","SaaS","DB","Medium","SDE,Backend","20-50 LPA","Remote","No","2+ yrs","turso.tech","Edge SQLite","74"),
  ("Appsmith","SaaS","Low-code","High","SDE,Backend","18-42 LPA","Bengaluru/Remote","Yes","0-4 yrs","appsmith.com/careers","Low-code builder","80"),
  ("Tooljet","SaaS","Low-code","High","SDE,Backend","16-38 LPA","Bengaluru/Remote","Yes","0-4 yrs","tooljet.com/careers","OSS low-code","78"),
  ("Retool","SaaS","Low-code","High","SDE,Backend","22-58 LPA","Remote","No","2+ yrs","retool.com/careers","Internal tools","82"),
  ("Vercel","SaaS","Frontend-Cloud","High","SDE,Backend,ML Eng","25-65 LPA","Remote","No","2+ yrs","vercel.com/careers","Next.js + AI","85"),
  ("Netlify","SaaS","Frontend-Cloud","Medium","SDE,Backend","20-50 LPA","Remote","No","2+ yrs","netlify.com/careers","Composable web","76"),
  ("Render","SaaS","Cloud","High","SDE,Backend,DevOps","22-55 LPA","Remote","No","2+ yrs","render.com/careers","Cloud platform","79"),
  ("Railway","SaaS","Cloud","Medium","SDE,Backend","20-48 LPA","Remote","No","2+ yrs","railway.app/careers","Dev cloud","73"),
  ("Clerk","SaaS","Auth","High","SDE,Backend","22-55 LPA","Remote","No","2+ yrs","clerk.com/careers","Auth-as-service","78"),
  ("Auth0/Okta","SaaS","Auth-Security","High","SDE,ML Eng,Backend","22-58 LPA","Remote/Bengaluru","No","2+ yrs","okta.com/company/careers","IAM expansion","84"),
  ("Linear","SaaS","ProjectMgmt","High","SDE,Backend","25-65 LPA","Remote","No","2+ yrs","linear.app/careers","PM platform","80"),
  # ── SERIES A/B STARTUPS ────────────────────────────────────────────────
  ("Kodo","Series A","Fintech","Medium","SDE,Backend,Data Analyst","12-28 LPA","Bengaluru","Yes","0-3 yrs","kodo.co.in/careers","Corporate cards","72"),
  ("Jar","Series B","Fintech","High","Backend,SDE,ML Eng","14-36 LPA","Bengaluru","Yes","0-4 yrs","jar.com/careers","Gold savings app","80"),
  ("Stashfin","Series B","Fintech","Medium","Backend,SDE,Data Sci","12-28 LPA","Delhi NCR","Yes","0-3 yrs","stashfin.com/careers","Personal loans","73"),
  ("Okcredit","Series B","Fintech","Medium","SDE,Backend","12-26 LPA","Bengaluru","Yes","0-3 yrs","okcredit.in/careers","SME credit","70"),
  ("Niyo","Series C","Neobank","High","Backend,SDE,Data Sci","14-34 LPA","Bengaluru","Yes","0-4 yrs","niyo.co/careers","Travel banking","78"),
  ("Smallcase","Series B","Fintech","High","Backend,SDE,Data Sci","14-36 LPA","Bengaluru","Yes","0-4 yrs","smallcase.com/careers","Investment platform","79"),
  ("Stockal","Series A","Fintech","Low","SDE,Data Analyst","10-22 LPA","Bengaluru","Yes","0-3 yrs","stockal.com","US stocks investing","55"),
  ("Vested Finance","Series A","Fintech","Low","SDE,Data Sci","10-22 LPA","Mumbai","Yes","0-3 yrs","vestedfinance.com/careers","US investing","54"),
  ("Freo","Series B","Neobank","Medium","Backend,SDE","12-26 LPA","Bengaluru","Yes","0-3 yrs","freo.money/careers","Credit line","68"),
  ("Axio","Series C","BNPL","Medium","Backend,SDE,Data Sci","14-32 LPA","Bengaluru","Yes","0-3 yrs","axio.in/careers","Buy now pay later","72"),
  ("Kredivo","Series C","BNPL","Medium","SDE,ML Eng","14-32 LPA","Bengaluru","Yes","0-3 yrs","kredivo.com/careers","SEA BNPL","70"),
  ("KreditBee","Series D","Fintech","High","Backend,SDE,ML Eng","14-34 LPA","Bengaluru","Yes","0-4 yrs","kreditbee.in/careers","Instant credit","79"),
  ("EarlySalary","Series C","Fintech","Medium","Backend,SDE","12-28 LPA","Pune","Yes","0-3 yrs","earlysalary.com/careers","Salary advance","70"),
  ("Moneyview","Series E","Fintech","High","Backend,SDE,ML Eng","14-36 LPA","Bengaluru","Yes","0-4 yrs","moneyview.in/careers","Personal finance AI","80"),
  ("OneCard","Series D","Fintech","High","Backend,SDE,ML Eng","14-36 LPA","Pune","Yes","0-4 yrs","getonecard.app/careers","Credit card tech","81"),
  # ── MORE COMPANIES ─────────────────────────────────────────────────────
  ("Juspay","Fintech","Payments","High","Backend,SDE,Data Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","juspay.in/careers","Payment orchestration","82"),
  ("Cashfree","Fintech","Payments","High","Backend,SDE,Data Eng","14-36 LPA","Bengaluru","Yes","0-4 yrs","cashfree.com/careers","B2B payments","82"),
  ("Safexpay","Fintech","Payments","Medium","Backend,SDE","10-24 LPA","Mumbai","Yes","0-3 yrs","safexpay.com/careers","Payment gateway","67"),
  ("Plural","Fintech","Payments","Medium","Backend,SDE","12-28 LPA","Bengaluru","Yes","0-3 yrs","pluralonline.com","UPI 2.0 infra","70"),
  ("PayU","Fintech","Payments","High","Backend,SDE,ML Eng","14-36 LPA","Bengaluru,Gurugram","Yes","0-4 yrs","payu.com/en-us/join-us","Prosus fintech","81"),
  ("Adyen","Fintech","Payments","High","SDE,Backend","22-58 LPA","Bengaluru","No","2+ yrs","adyen.com/careers","Global payments","83"),
  ("Truecaller","Swedish Product","CommApp","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru","Yes","0-4 yrs","truecaller.com/en/about/jobs","Spam-AI expansion","83"),
  ("ShareChat","Indian Product","Social","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru","Yes","0-4 yrs","sharechat.com/careers","Vernacular AI","83"),
  ("Dailyhunt","Indian Product","Content","Medium","SDE,ML Eng,Data Sci","14-34 LPA","Bengaluru","Yes","0-3 yrs","dailyhunt.in/careers","News AI","73"),
  ("Josh (DH)","Indian Product","Short-video","Medium","SDE,ML Eng","14-34 LPA","Bengaluru","Yes","0-3 yrs","dailyhunt.in/careers","Reels competitor","73"),
  ("Moj (Sharechat)","Indian Product","Short-video","Medium","SDE,ML Eng","14-34 LPA","Bengaluru","Yes","0-3 yrs","sharechat.com/careers","Short video AI","73"),
  ("Wynk Music","Indian Product","Music","Low","SDE,Backend","10-24 LPA","Delhi NCR","Yes","0-3 yrs","airtel.in/careers","Streaming platform","58"),
  ("Gaana","Indian Product","Music","Low","SDE,Backend","10-22 LPA","Bengaluru","Yes","0-3 yrs","gaana.com/about-us/jobs","Music streaming","54"),
  ("Jio Platform","Indian Product","Telecom-Tech","Very High","SDE,ML Eng,Data Sci","12-35 LPA","Mumbai,Hyderabad","Yes","0-5 yrs","jio.com/careers","Jio 5G + AI","90"),
  ("Airtel","Indian Product","Telecom","High","SDE,Backend,Data Sci","12-32 LPA","Delhi NCR,Bengaluru","Yes","0-4 yrs","airtel.in/careers","Digital services","84"),
  ("BYJU's","EdTech","EdTech","Low","SDE,Data Analyst","8-20 LPA","Bengaluru","Yes","0-3 yrs","byjus.com/careers","Restructuring phase","40"),
  ("Unacademy","EdTech","EdTech","Medium","SDE,ML Eng,Data Sci","12-28 LPA","Bengaluru","Yes","0-3 yrs","unacademy.com/careers","AI tutoring relaunch","68"),
  ("Classplus","EdTech","EdTech","Medium","SDE,Backend","10-24 LPA","Delhi NCR","Yes","0-3 yrs","classplus.co/careers","Coaching platform","69"),
  ("Topper (Toppr)","EdTech","EdTech","Low","SDE,Data Analyst","8-20 LPA","Mumbai","Yes","0-3 yrs","toppr.com/careers","Platform pivot","50"),
  ("Vedantu","EdTech","EdTech","Medium","SDE,ML Eng","12-28 LPA","Bengaluru","Yes","0-3 yrs","vedantu.com/careers","Live classes AI","68"),
  ("Simplilearn","EdTech","EdTech","High","SDE,Data Sci,ML Eng","12-32 LPA","Bengaluru","Yes","0-4 yrs","simplilearn.com/careers","UpSkilling AI","78"),
  ("Great Learning","EdTech","EdTech","High","SDE,Data Sci,ML Eng","12-30 LPA","Hyderabad","Yes","0-4 yrs","greatlearning.in/careers","AI education","78"),
  ("Naukri (Infoedge)","Indian Product","HRTech","High","SDE,ML Eng,Data Sci","14-38 LPA","Noida","Yes","0-4 yrs","infoedge.in/careers","AI job matching","84"),
  ("Shine.com","Indian Product","HRTech","Medium","SDE,Data Analyst","10-24 LPA","Delhi NCR","Yes","0-3 yrs","shine.com/careers","Job portal","67"),
  ("Hirist.tech","Indian Product","HRTech","Medium","SDE,Data Analyst","10-22 LPA","Delhi NCR","Yes","0-3 yrs","hirist.tech","Tech hiring","65"),
  ("Apna","Indian Product","HRTech","High","SDE,ML Eng,Backend","14-34 LPA","Bengaluru","Yes","0-4 yrs","apna.co/careers","Blue-collar jobs AI","80"),
  ("WorkIndia","Indian Product","HRTech","Medium","SDE,Backend","10-24 LPA","Bengaluru","Yes","0-3 yrs","workindia.in/careers","Tier-2 hiring","68"),
  ("Springworks","Indian Product","HRTech","Medium","SDE,Backend","12-28 LPA","Bengaluru","Yes","0-3 yrs","springworks.in/careers","HR tools","68"),
  ("Keka HR","Indian Product","HRTech","High","SDE,Backend,Data Sci","14-34 LPA","Hyderabad","Yes","0-4 yrs","keka.com/careers","HR SaaS growth","79"),
  ("GreytHR","Indian Product","HRTech","Medium","SDE,Backend","10-24 LPA","Bengaluru","Yes","0-3 yrs","greythr.com/careers","Payroll SaaS","67"),
  ("HRMantra","Indian Product","HRTech","Low","SDE,Backend","8-20 LPA","Mumbai","Yes","0-3 yrs","hrmantra.com/careers","HRMS","55"),
  ("Zeta","Fintech","Banking-Infra","High","Backend,SDE,ML Eng","16-42 LPA","Bengaluru","Yes","0-4 yrs","zeta.tech/careers","Neo-banking infra","84"),
  ("Epifi","Neobank","Banking","High","Backend,SDE,Data Sci","16-40 LPA","Bengaluru","Yes","0-4 yrs","fi.money/careers","Open banking","81"),
  ("Yubi (CredAvenue)","Fintech","Debt-Market","High","Backend,SDE,Data Sci","14-38 LPA","Chennai","Yes","0-4 yrs","yubi.co/careers","Debt platform","80"),
  ("Indifi","Fintech","SME Lending","Medium","SDE,Backend,Data Sci","12-28 LPA","Gurugram","Yes","0-3 yrs","indifi.com/careers","SME credit","70"),
  ("Mintifi","Fintech","SCF","Medium","SDE,Backend","12-26 LPA","Mumbai","Yes","0-3 yrs","mintifi.com/careers","Supply chain finance","68"),
  ("NeoGrowth","Fintech","MSME Lending","Medium","SDE,Data Sci","10-24 LPA","Mumbai","Yes","0-3 yrs","neogrowth.in/careers","Cash flow lending","66"),
  ("U GRO Capital","Fintech","MSME Lending","Medium","SDE,Data Sci","10-22 LPA","Mumbai","Yes","0-3 yrs","ugrocapital.com/careers","Sector lending","64"),
  ("Nucleus Software","Fintech","Banking-SW","Medium","SDE,Backend","10-28 LPA","Noida","Yes","0-4 yrs","nucleussoftware.com/careers","Loan origination","69"),
  ("Finastra","GCC","Banking-SW","High","SDE,Backend,Data Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","finastra.com/careers","Open finance","80"),
  ("FIS Global","GCC","Fintech","High","SDE,ML Eng,Data Sci","14-38 LPA","Bengaluru,Pune","Yes","0-4 yrs","fisglobal.com/en/about/careers","Payments AI","80"),
  ("Fiserv","GCC","Fintech","High","SDE,Backend,ML Eng","14-38 LPA","Bengaluru,Pune","Yes","0-4 yrs","fiserv.com/en/about-fiserv/careers","Banking tech","80"),
  ("Jack Henry","GCC","Banking-SW","Medium","SDE,Backend","14-32 LPA","Remote","Yes","0-3 yrs","jackhenry.com/working-at-jack-henry","Community banking","73"),
  ("Temenos","GCC","Banking-SW","Medium","SDE,Backend,Data Eng","14-34 LPA","Bengaluru","Yes","0-3 yrs","temenos.com/careers","Core banking","73"),
  ("Calypso Technology","GCC","Capital Markets","Medium","SDE,Backend","16-38 LPA","Bengaluru","No","2+ yrs","finastra.com/careers","Treasury tech","72"),
  ("Murex","GCC","Capital Markets","Medium","SDE,Backend","16-38 LPA","Bengaluru","No","2+ yrs","murex.com/careers","Trading platforms","72"),
  ("ION Group","GCC","Capital Markets","Medium","SDE,Backend,Data Sci","14-36 LPA","Bengaluru","No","2+ yrs","iongroup.com/careers","Financial tech","71"),
  ("FIS-Metavante","GCC","Payments","Medium","SDE,Backend","12-30 LPA","Bengaluru","Yes","0-3 yrs","fisglobal.com","Payment rails","69"),
  ("Broadridge","GCC","FinOps","High","SDE,Data Sci,ML Eng","16-40 LPA","Hyderabad","Yes","0-4 yrs","broadridge.com/careers","AI fintech","80"),
  ("SS&C Technologies","GCC","FinTech","High","SDE,Data Eng,BI Analyst","14-36 LPA","Hyderabad,Mumbai","Yes","0-4 yrs","ssctech.com/careers","Fund operations AI","79"),
  ("Virtusa","Service","IT-Fintech","High","SDE,ML Eng,Data Eng","10-30 LPA","Hyderabad,Chennai","Yes","0-4 yrs","virtusa.com/careers","Digital engineering","78"),
  ("UST Global","Service","IT","High","SDE,ML Eng,Backend","10-28 LPA","Thiruvananthapuram,Bengaluru","Yes","0-4 yrs","ust.com/en/careers","Cloud + AI","77"),
  ("EPAM Systems","Service","Digital-Eng","High","SDE,ML Eng,Data Sci","14-40 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","epam.com/careers","AI solutions","83"),
  ("GlobalLogic","Service","Digital-Eng","High","SDE,ML Eng,Backend","14-40 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","globallogic.com/careers","Hitachi Group","82"),
  ("Nagarro","Service","Digital-Eng","High","SDE,ML Eng,Data Sci","12-34 LPA","Gurugram,Bengaluru","Yes","0-4 yrs","nagarro.com/en/careers","Digital product eng","80"),
  ("Thoughtworks","Service","Digital-Eng","High","SDE,ML Eng,Data Sci","14-40 LPA","Bengaluru,Pune","Yes","0-4 yrs","thoughtworks.com/careers","AI practices","82"),
  ("Publicis Sapient","Service","Digital","High","SDE,ML Eng,Data Sci","12-35 LPA","Bengaluru,Gurugram","Yes","0-4 yrs","sapient.com/careers","GenAI services","80"),
  ("Harman Digital","GCC","Automotive-Tech","High","SDE,ML Eng,Data Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","harman.com/connect/careers","Samsung GCC","82"),
  ("Samsung R&D","GCC","Consumer-Tech","Very High","SDE,ML Eng,Data Sci","16-45 LPA","Noida,Bengaluru","Yes","0-5 yrs","samsung.com/in/aboutsamsung/careers","Galaxy AI","89"),
  ("Motorola Solutions","GCC","Safety-Tech","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","motorolasolutions.com/en_us/careers","Public safety AI","83"),
  ("Dell Technologies","GCC","IT-Infra","High","SDE,ML Eng,Data Sci","14-40 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","dell.com/en-us/dt/corporate/careers","AI PC + infra","83"),
  ("HP Inc","GCC","IT","High","SDE,ML Eng,Data Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","hp.com/h20195/v2/getpdf.aspx/careers","AI compute","81"),
  ("HPE","GCC","IT-Infra","High","SDE,ML Eng,Data Sci","16-42 LPA","Bengaluru,Pune","Yes","0-4 yrs","hpe.com/h41271/404d.aspx?0x3e0000","GreenLake AI","82"),
  ("Lenovo","GCC","IT","Medium","SDE,Data Analyst","12-30 LPA","Bengaluru","Yes","0-3 yrs","lenovocareers.com","AI devices","74"),
  ("Cisco","GCC","Networking","Very High","SDE,ML Eng,Data Sci","18-50 LPA","Bengaluru","Yes","0-5 yrs","cisco.com/c/en/us/about/careers","AI networking","89"),
  ("Juniper Networks","GCC","Networking","High","SDE,ML Eng,Backend","16-44 LPA","Bengaluru","Yes","0-4 yrs","juniper.net/us/en/company/careers","AI-Native networking","84"),
  ("Arista Networks","GCC","Networking","High","SDE,Backend","18-50 LPA","Bengaluru","No","2+ yrs","arista.com/en/company/careers","Cloud networking","83"),
  ("F5 Networks","GCC","App Security","High","SDE,ML Eng,Data Sci","18-48 LPA","Hyderabad","No","2+ yrs","f5.com/company/careers","NGINX + AI security","82"),
  ("Akamai","GCC","CDN-Security","High","SDE,ML Eng,Data Eng","18-48 LPA","Bengaluru","No","2+ yrs","akamai.com/company/careers","Edge security","82"),
  ("Rackspace","Service","Cloud","Medium","SDE,DevOps,Data Eng","12-30 LPA","Bengaluru","Yes","0-4 yrs","rackspace.com/talent","Multi-cloud services","72"),
  ("DXC Technology","Service","IT","High","SDE,ML Eng,Data Sci","10-30 LPA","Bengaluru,Chennai","Yes","0-4 yrs","dxc.com/us/en/careers","DX+AI","78"),
  ("NIIT Digital","Service","IT","Medium","SDE,Data Analyst","8-20 LPA","Delhi NCR","Yes","0-3 yrs","niit.com/careers","SkillTech","66"),
  ("Kyndryl","Service","IT-Infra","High","SDE,ML Eng,Data Sci","14-36 LPA","Bengaluru,Chennai","Yes","0-4 yrs","kyndryl.com/us/en/careers","IBM spin-off","78"),
  ("NTT Data","Service","IT","High","SDE,ML Eng,Data Sci","12-32 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","nttdata.com/global/en/careers","Digital engineering","78"),
  ("Fujitsu India","Service","IT","High","SDE,ML Eng,Data Eng","12-32 LPA","Noida,Bengaluru","Yes","0-4 yrs","fujitsu.com/global/about/careers","AI transformation","77"),
  ("CGI Group","Service","IT","High","SDE,Backend,Data Sci","12-32 LPA","Bengaluru,Hyderabad","Yes","0-4 yrs","cgi.com/en/careers","Govtech AI","77"),
  ("Atos","Service","IT","Medium","SDE,Data Sci","10-28 LPA","Bengaluru,Pune","Yes","0-4 yrs","atos.net/en/careers","Cloud services","70"),
  ("Unisys","Service","IT","Medium","SDE,Data Eng,ML Eng","12-30 LPA","Bengaluru","Yes","0-4 yrs","unisys.com/careers","Logistics AI","72"),
  ("L&T Technology Services","Service","Engg","High","SDE,ML Eng,Data Sci","12-34 LPA","Vadodara,Bengaluru","Yes","0-4 yrs","ltts.com/careers","Industrial IoT","80"),
  ("Coforge","Service","IT","High","SDE,ML Eng,Backend","10-30 LPA","Noida","Yes","0-4 yrs","coforge.com/careers","BFSI + Travel","78"),
  ("Hexaware","Service","IT","High","SDE,ML Eng,Data Sci","10-30 LPA","Chennai,Mumbai","Yes","0-4 yrs","hexaware.com/careers","BPS + AI","79"),
  ("Mastech Holdings","Service","Staffing","Medium","SDE,Data Sci","10-24 LPA","Bengaluru","Yes","0-3 yrs","mastech.com/careers","IT staffing","65"),
  ("Xoriant","Service","Digital-Eng","Medium","SDE,ML Eng","12-30 LPA","Pune","Yes","0-3 yrs","xoriant.com/careers","Product engineering","69"),
  ("Kellton Tech","Service","IT","Medium","SDE,Backend","10-26 LPA","Delhi NCR,Hyderabad","Yes","0-3 yrs","kelltontech.com/careers","Digital solutions","66"),
  # ── YC / SEED / EARLY STARTUPS ─────────────────────────────────────────
  ("Replit","AI Startup","Dev-AI","High","SDE,ML Eng,Backend","28-70 LPA","Remote","No","2+ yrs","replit.com/site/careers","AI coding env","84"),
  ("Bolt.new","AI Startup","Dev-AI","High","SDE,ML Eng,Backend","28-72 LPA","Remote","No","2+ yrs","stackblitz.com/careers","Full-stack AI","84"),
  ("Lovable","AI Startup","Dev-AI","High","SDE,ML Eng","28-70 LPA","Remote","No","2+ yrs","lovable.dev/careers","AI product builder","82"),
  ("Dify","AI Startup","LLM-Ops","High","SDE,ML Eng,Backend","22-58 LPA","Remote","No","2+ yrs","dify.ai","LLM app platform","82"),
  ("LangChain","AI Startup","LLM-Infra","High","SDE,ML Eng","25-65 LPA","Remote","No","2+ yrs","langchain.com/careers","LLM orchestration","84"),
  ("LlamaIndex","AI Startup","RAG","High","SDE,ML Eng,Data Sci","25-65 LPA","Remote","No","2+ yrs","llamaindex.ai/careers","RAG framework","82"),
  ("Fixie.ai","AI Startup","AgentAI","Medium","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","fixie.ai","AI agents","76"),
  ("AutoGPT","AI Startup","AgentAI","Medium","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","agpt.co","OSS AI agents","74"),
  ("Voyage AI","AI Startup","Embeddings","Medium","ML Eng,SDE","25-65 LPA","Remote","No","2+ yrs","voyageai.com","Embedding models","76"),
  ("Weaviate","AI Startup","VectorDB","High","SDE,ML Eng,Backend","25-65 LPA","Remote","No","2+ yrs","weaviate.io/company/careers","Vector search","82"),
  ("Pinecone","AI Startup","VectorDB","High","SDE,ML Eng,Backend","28-72 LPA","Remote","No","2+ yrs","pinecone.io/company/careers","Serverless vector","83"),
  ("Qdrant","AI Startup","VectorDB","Medium","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","qdrant.tech","OSS vector DB","76"),
  ("Chroma","AI Startup","VectorDB","Medium","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","trychroma.com","Embedding DB","74"),
  ("Activeloop","AI Startup","Data-Infra","Medium","SDE,ML Eng,Data Eng","22-52 LPA","Remote","No","2+ yrs","activeloop.ai","Deep Lake","73"),
  ("Scale AI","AI Startup","Data-Label","Very High","ML Eng,Data Sci,SDE","25-70 LPA","Remote","No","2+ yrs","scale.com/careers","RLHF data","88"),
  ("Labelbox","AI Startup","Data-Label","High","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","labelbox.com/careers","Training data","80"),
  ("Cleanlab","AI Startup","Data-Quality","Medium","ML Eng,Data Sci","22-55 LPA","Remote","No","2+ yrs","cleanlab.ai/careers","Data quality AI","76"),
  ("Encord","AI Startup","Vision-AI","Medium","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","encord.com/careers","Video labeling","75"),
  ("Galileo","AI Startup","LLMOps","High","SDE,ML Eng,Data Sci","25-65 LPA","Remote","No","2+ yrs","rungalileo.io/careers","LLM evaluation","80"),
  ("Confident AI","AI Startup","LLMOps","Medium","SDE,ML Eng","22-52 LPA","Remote","No","2+ yrs","confident-ai.com","DeepEval","73"),
  ("Arize AI","AI Startup","MLOps","High","SDE,MLOps,ML Eng","25-65 LPA","Remote","No","2+ yrs","arize.com/company/careers","Model observability","82"),
  ("Evidently AI","AI Startup","MLOps","High","SDE,MLOps,ML Eng","22-58 LPA","Remote","No","2+ yrs","evidentlyai.com/careers","ML monitoring","80"),
  ("Fiddler AI","AI Startup","MLOps","High","SDE,ML Eng,MLOps","22-58 LPA","Remote/Bengaluru","No","2+ yrs","fiddler.ai/careers","AI observability","80"),
  ("Seldon","AI Startup","MLOps","Medium","SDE,ML Eng,MLOps","22-52 LPA","Remote","No","2+ yrs","seldon.io/careers","Model deployment","74"),
  ("ClearML","AI Startup","MLOps","Medium","SDE,ML Eng,MLOps","20-50 LPA","Remote","No","2+ yrs","clearml.com/careers","OSS MLOps","73"),
  ("Bentoml","AI Startup","MLOps","Medium","SDE,ML Eng","22-55 LPA","Remote","No","2+ yrs","bentoml.com/careers","Model serving","74"),
  ("Replicate","AI Startup","MLOps","High","SDE,ML Eng,Backend","28-72 LPA","Remote","No","2+ yrs","replicate.com/careers","OSS model hosting","82"),
  ("Hugging Face","AI Company","LLM-Hub","Very High","ML Eng,SDE,Data Sci","28-75 LPA","Remote/Paris","No","2+ yrs","huggingface.co/jobs","OSS AI hub","90"),
  ("Lightning AI","AI Startup","MLOps","High","SDE,ML Eng,MLOps","25-65 LPA","Remote","No","2+ yrs","lightning.ai/careers","PyTorch Lightning","82"),
  ("Anyscale","AI Startup","Distributed-AI","High","SDE,ML Eng,MLOps","28-72 LPA","Remote","No","2+ yrs","anyscale.com/careers","Ray framework","84"),
  ("Determined AI","AI Startup","MLOps","Medium","SDE,ML Eng,MLOps","25-62 LPA","Remote","No","2+ yrs","hpe.com","HPE ML platform","76"),
  ("ZenML","AI Startup","MLOps","Medium","SDE,ML Eng,MLOps","20-50 LPA","Remote","No","2+ yrs","zenml.io/careers","OSS MLOps","73"),
  ("Metaflow","AI Startup","MLOps","Medium","SDE,ML Eng","20-50 LPA","Remote","No","2+ yrs","netflix.com/jobs","Netflix MLOps","74"),
  ("Prefect","AI Startup","Orchestration","High","SDE,Data Eng,MLOps","22-55 LPA","Remote","No","2+ yrs","prefect.io/careers","Data/ML workflows","78"),
  ("Dagster","AI Startup","Orchestration","High","SDE,Data Eng,MLOps","22-55 LPA","Remote","No","2+ yrs","dagster.io/careers","Data orchestration","78"),
  ("dbt Labs","SaaS","Data-Transform","Very High","Data Eng,Analytics Eng","22-58 LPA","Remote","No","2+ yrs","getdbt.com/dbt-labs/careers","Analytics engineer","87"),
  ("Fivetran","SaaS","ELT","High","Data Eng,SDE,Backend","22-58 LPA","Remote/Bengaluru","No","2+ yrs","fivetran.com/careers","Data integration","84"),
  ("Monte Carlo","SaaS","Data-Observ","High","SDE,Data Eng,ML Eng","25-65 LPA","Remote","No","2+ yrs","montecarlodata.com/careers","Data reliability","80"),
  ("Atlan","Indian SaaS","Data-Catalog","Very High","SDE,Data Eng,ML Eng","18-48 LPA","Delhi NCR","Yes","0-4 yrs","atlan.com/careers","Metadata AI","88"),
  ("Lightdash","SaaS","BI","Medium","SDE,Data Analyst","20-48 LPA","Remote","No","2+ yrs","lightdash.com/careers","BI for dbt","72"),
  ("Hex Tech","SaaS","Analytics","High","SDE,Data Sci,Data Analyst","22-55 LPA","Remote","No","2+ yrs","hex.tech/careers","Notebooks + AI","79"),
  ("Observable","SaaS","Data Viz","Medium","SDE,Data Sci","20-50 LPA","Remote","No","2+ yrs","observablehq.com/jobs","D3 + AI","72"),
  ("Streamlit","SaaS","Data Apps","High","SDE,ML Eng,Data Sci","22-55 LPA","Remote/Snowflake","No","2+ yrs","snowflake.com/en/company/careers","Snowflake owned","78"),
  ("Gradio","SaaS","ML-UI","Medium","SDE,ML Eng","20-50 LPA","Remote/HuggingFace","No","2+ yrs","huggingface.co/jobs","HuggingFace owned","74"),
  ("Plotly/Dash","SaaS","Data Viz","Medium","SDE,Data Sci","18-45 LPA","Remote","No","2+ yrs","plotly.com/company/careers","Enterprise analytics","72"),
  ("Tableau/Salesforce","GCC/SaaS","BI","High","SDE,Data Analyst,BI Analyst","20-50 LPA","Bengaluru","No","2+ yrs","salesforce.com/careers","Salesforce owned","80"),
  ("Power BI/Microsoft","GCC/SaaS","BI","High","SDE,BI Analyst,Data Sci","18-48 LPA","Hyderabad","Yes","0-4 yrs","careers.microsoft.com","Microsoft Fabric","82"),
  ("Looker/Google","GCC/SaaS","BI","High","SDE,Data Analyst,BI Analyst","20-52 LPA","Bengaluru","No","2+ yrs","careers.google.com","Google Cloud BI","82"),
  ("ThoughtSpot","SaaS","BI-AI","High","SDE,ML Eng,Data Sci","20-52 LPA","Bengaluru","No","2+ yrs","thoughtspot.com/company/careers","AI analytics","81"),
  ("Sigma Computing","SaaS","BI","Medium","SDE,Data Analyst","20-50 LPA","Remote","No","2+ yrs","sigmacomputing.com/careers","Cloud BI","74"),
  ("Sisense","SaaS","BI","Medium","SDE,Data Analyst,BI Analyst","18-44 LPA","Remote","No","2+ yrs","sisense.com/company/careers","Embedded analytics","72"),
  ("Qlik","SaaS","BI","Medium","SDE,Data Analyst,BI Analyst","18-44 LPA","Remote/Bengaluru","No","2+ yrs","qlik.com/us/company/careers","Augmented analytics","73"),
  ("MicroStrategy","SaaS","BI","Low","SDE,Data Analyst","16-36 LPA","Remote","No","2+ yrs","microstrategy.com/en/jobs","Bitcoin BI pivot","60"),
  ("Yellowbrick","SaaS","Data-Warehouse","Low","SDE,Data Eng","18-42 LPA","Remote","No","2+ yrs","yellowbrick.com/company/careers","Cloud DW","62"),
  ("Starburst","SaaS","Query-Engine","Medium","SDE,Data Eng","22-55 LPA","Remote","No","2+ yrs","starburst.io/careers","Trino-based","74"),
  ("Dremio","SaaS","Data-Lakehouse","Medium","SDE,Data Eng","22-55 LPA","Remote","No","2+ yrs","dremio.com/company/careers","Lakehouse","74"),
  ("Cloudera","SaaS","Data-Platform","Medium","SDE,Data Eng,ML Eng","16-42 LPA","Remote/Bengaluru","Yes","0-4 yrs","cloudera.com/careers","CDP platform","73"),
  ("Palantir","Product","AI-Platforms","High","SDE,ML Eng,Data Sci","30-80 LPA","Bengaluru/Remote","No","3+ yrs","palantir.com/careers","AI platform","87"),
  ("C3.ai","AI Company","Enterprise-AI","Medium","SDE,ML Eng,Data Sci","25-65 LPA","Remote","No","2+ yrs","c3.ai/company/careers","Enterprise AI SaaS","77"),
  ("DataRobot","AI Company","AutoML","Medium","SDE,ML Eng,Data Sci","22-58 LPA","Remote","No","2+ yrs","datarobot.com/company/careers","AutoML platform","76"),
  ("H2O.ai","AI Company","AutoML","Medium","SDE,ML Eng,Data Sci","22-55 LPA","Remote","No","2+ yrs","h2o.ai/company/careers","OSS AutoML","74"),
  ("BenevolentAI","AI Company","BioAI","Medium","ML Eng,Data Sci","25-65 LPA","Remote","No","3+ yrs","benevolent.ai/careers","Drug discovery AI","76"),
  ("Recursion","AI Company","BioAI","Medium","ML Eng,Data Sci","25-65 LPA","Remote","No","3+ yrs","recursion.com/careers","Bio AI","76"),
  ("Insilico Medicine","AI Company","BioAI","Medium","ML Eng,Data Sci","22-60 LPA","Remote","No","3+ yrs","insilico.com/careers","Drug AI","74"),
  ("Tempus AI","AI Company","MedAI","Medium","ML Eng,Data Sci","25-65 LPA","Remote","No","3+ yrs","tempus.com/careers","Clinical AI","76"),
  ("PathAI","AI Company","MedAI","Medium","ML Eng,Data Sci","25-65 LPA","Remote","No","3+ yrs","pathai.com/company/careers","Pathology AI","75"),
  ("Babylon Health","AI Startup","HealthAI","Low","ML Eng,SDE","18-42 LPA","Remote","No","2+ yrs","babylonhealth.com/careers","Health AI","60"),
  ("Practo","Indian Product","HealthTech","High","SDE,ML Eng,Data Sci","14-36 LPA","Bengaluru","Yes","0-4 yrs","practo.com/company/careers","Digital health","80"),
  ("1mg (Tata Health)","Indian Product","HealthTech","High","SDE,ML Eng,Data Sci","14-36 LPA","Gurugram","Yes","0-4 yrs","1mg.com/careers","Pharma AI","80"),
  ("PharmEasy","Indian Product","HealthTech","Medium","SDE,Data Sci","12-28 LPA","Mumbai","Yes","0-3 yrs","pharmeasy.in/careers","E-pharmacy","71"),
  ("Docprime","Indian Product","HealthTech","Low","SDE,Data Analyst","10-22 LPA","Delhi NCR","Yes","0-3 yrs","docprime.com","Doctor booking","55"),
  ("Portea Medical","Indian Product","HealthTech","Medium","SDE,Data Analyst","10-24 LPA","Bengaluru","Yes","0-3 yrs","portea.com/careers","Home healthcare","64"),
  ("Mfine","Indian Product","HealthAI","Medium","ML Eng,SDE","14-32 LPA","Bengaluru","Yes","0-3 yrs","mfine.co/careers","AI diagnostics","72"),
  ("Eka Care","AI Startup","HealthAI","High","SDE,ML Eng,Data Sci","16-38 LPA","Bengaluru","Yes","0-4 yrs","eka.care/careers","PHR AI","78"),
  ("Curefit","Indian Product","HealthFit","Medium","SDE,Data Sci","12-28 LPA","Bengaluru","Yes","0-3 yrs","cure.fit/careers","Health AI","70"),
  ("Nykaa Health","Indian Product","Wellness","Medium","SDE,Backend","12-26 LPA","Mumbai","Yes","0-3 yrs","nykaa.com/careers","Wellness platform","68"),
  ("Ola Electric","Indian Product","EV","High","SDE,ML Eng,Data Eng","14-38 LPA","Bengaluru","Yes","0-4 yrs","olaelectric.com/careers","EV software","82"),
  ("Ather Energy","Indian Product","EV","High","SDE,ML Eng,Data Sci","14-38 LPA","Bengaluru,Chennai","Yes","0-4 yrs","atherenergy.com/careers","EV AI","82"),
  ("Tata Motors Digital","GCC","Automotive","High","SDE,ML Eng,Data Sci","14-38 LPA","Pune","Yes","0-4 yrs","tatamotors.com/careers","TPEM software","81"),
  ("Hero Electric","Indian Product","EV","Medium","SDE,Backend","12-26 LPA","Delhi NCR","Yes","0-3 yrs","heroelectric.in/careers","2W EV","68"),
  ("BluSmart","Indian Product","EV-Mobility","High","Backend,SDE,Data Sci","14-34 LPA","Delhi NCR,Bengaluru","Yes","0-4 yrs","blusmart.in/careers","EV fleet tech","78"),
  ("Yulu","Indian Product","EV-Micro","Medium","SDE,Backend","12-26 LPA","Bengaluru","Yes","0-3 yrs","yulu.bike/careers","Micro-mobility","68"),
  ("Log9 Materials","Indian Product","EV-Battery","Medium","SDE,Data Sci","12-28 LPA","Bengaluru","Yes","0-3 yrs","log9materials.com/careers","Battery AI","68"),
  ("Ultraviolette","Indian Product","EV","Medium","SDE,ML Eng","12-28 LPA","Bengaluru","Yes","0-3 yrs","ultraviolette.com/careers","Electric motorcycle","68"),
  ("IndiGrid","Indian Product","Energy","Low","SDE,Data Analyst","10-22 LPA","Mumbai","Yes","0-3 yrs","indigrid.com/careers","Grid tech","55"),
  ("Greenko","Indian Product","CleanEnergy","Medium","SDE,Data Eng","10-24 LPA","Hyderabad","Yes","0-3 yrs","greenko.in/careers","Renewable ops","63"),
  ("Avaada Group","Indian Product","CleanEnergy","Low","SDE,Data Analyst","10-20 LPA","Delhi NCR","Yes","0-3 yrs","avaada.com/careers","Solar operations","54"),
  ("ReNew Power","Indian Product","CleanEnergy","Medium","SDE,Data Eng","10-24 LPA","Gurugram","Yes","0-3 yrs","renewpower.in/careers","Wind+Solar data","63"),
  ("Tata Power Digital","GCC","Energy-Tech","High","SDE,ML Eng,Data Sci","14-36 LPA","Mumbai","Yes","0-4 yrs","tatapower.com/careers","Smart grid AI","79"),
  ("Adani Digital Labs","GCC","Conglomerate","High","SDE,ML Eng,Data Sci","14-38 LPA","Ahmedabad,Mumbai","Yes","0-4 yrs","adani.com/careers","Group tech hub","80"),
  ("Jio Platforms","Indian Product","Telecom-AI","Very High","SDE,ML Eng,Data Sci","14-38 LPA","Mumbai,Bengaluru,Hyderabad","Yes","0-5 yrs","jio.com/careers","India digital infra","91"),
  ("BPCL Digital","GCC","Energy","Medium","SDE,Data Analyst","10-24 LPA","Mumbai","Yes","0-3 yrs","bharatpetroleum.in","Oil+gas digital","63"),
  ("ONGC Digital","GCC","Energy","Medium","SDE,Data Sci","10-24 LPA","Delhi NCR","Yes","0-3 yrs","ongcindia.com","Energy data","62"),
  ("L&T Infotech (LTI)","Service","IT","High","SDE,ML Eng,Data Sci","12-34 LPA","Mumbai,Pune","Yes","0-5 yrs","ltimindtree.com/careers","Digital eng","80"),
  ("NIIT University","EdTech","EdTech","Low","SDE,Data Analyst","6-14 LPA","Delhi NCR","Yes","0-2 yrs","niituniversity.in","Academic","45"),
  ("Acko","Indian Product","InsurTech","High","SDE,ML Eng,Data Sci","16-40 LPA","Bengaluru","Yes","0-4 yrs","acko.com/careers","Digital insurance AI","82"),
  ("Digit Insurance","Indian Product","InsurTech","High","SDE,ML Eng,Data Sci","14-38 LPA","Bengaluru","Yes","0-4 yrs","godigit.com/careers","InsurAI","82"),
  ("PolicyBazaar","Indian Product","InsurTech","High","Backend,SDE,Data Sci","14-36 LPA","Gurugram","Yes","0-4 yrs","policybazaar.com/careers","Insurance AI","81"),
  ("Coverfox","Indian Product","InsurTech","Medium","SDE,Data Analyst","10-24 LPA","Mumbai","Yes","0-3 yrs","coverfox.com/careers","Insurance platform","68"),
  ("Turtlemint","Indian Product","InsurTech","Medium","Backend,SDE","12-28 LPA","Mumbai","Yes","0-3 yrs","turtlemint.com/careers","Distribution AI","70"),
  ("Riskcovry","Indian Product","InsurTech","Medium","SDE,Backend","12-26 LPA","Mumbai","Yes","0-3 yrs","riskcovry.com/careers","InsurTech API","67"),
  ("Flock","Indian Product","InsurTech","Medium","SDE,Backend","12-26 LPA","Mumbai","Yes","0-3 yrs","flock.com.co/careers","Commercial lines","66"),
  ("Onsurity","Indian Product","InsurTech","Medium","SDE,Backend","12-26 LPA","Bengaluru","Yes","0-3 yrs","onsurity.com/careers","Group health","67"),
  ("Plum Benefits","Indian Product","InsurTech","Medium","SDE,Backend","12-26 LPA","Bengaluru","Yes","0-3 yrs","plumhq.com/careers","Employee health","67"),
  ("Piramal Finance","Indian Product","NBFC","Medium","SDE,Data Sci,ML Eng","12-30 LPA","Mumbai","Yes","0-3 yrs","piramal.com/careers","Lending AI","70"),
  ("Mahindra Finance","GCC","NBFC","Medium","SDE,Data Analyst","10-24 LPA","Mumbai","Yes","0-3 yrs","mahindrafinance.com/careers","Rural fintech","62"),
  ("Bajaj Finserv","Indian Product","NBFC","High","SDE,ML Eng,Data Sci","14-38 LPA","Pune","Yes","0-4 yrs","bajajfinserv.in/careers","Digital lending AI","82"),
  ("HDFC Digital","GCC","Banking","High","SDE,ML Eng,Data Sci","14-38 LPA","Mumbai,Bengaluru","Yes","0-4 yrs","hdfcbank.com/content/bbp/repositories/723fb80a-2dde-42a3-9793-7ae1be57c87f/?folderName=/OtherSection/Careers/careers_at_hdfc","Retail banking AI","82"),
  ("ICICI Digital","GCC","Banking","High","SDE,ML Eng,Data Sci","14-38 LPA","Mumbai,Hyderabad","Yes","0-4 yrs","icicicareers.com","iMobile AI","82"),
  ("Axis Bank Digital","GCC","Banking","High","SDE,ML Eng,Data Sci","14-36 LPA","Mumbai","Yes","0-4 yrs","axisbank.com/about-us/careers","Digital banking","80"),
  ("SBI Digital","GCC","Banking","Medium","SDE,Data Analyst","10-26 LPA","Mumbai,Hyderabad","Yes","0-4 yrs","sbi.co.in/web/careers","YONO platform","72"),
  ("Kotak Mahindra Digital","GCC","Banking","High","SDE,ML Eng,Data Sci","14-36 LPA","Mumbai","Yes","0-4 yrs","kotak.com/en/about-us/careers","Kotak811 AI","79"),
  ("Zerodha Fund House","Indian Product","Asset Mgmt","Medium","SDE,Data Analyst","12-28 LPA","Bengaluru","Yes","0-3 yrs","zerodha.com/careers","AMC tech","69"),
  ("ET Money","Indian Product","Fintech","Medium","SDE,ML Eng","12-30 LPA","Delhi NCR","Yes","0-3 yrs","etmoney.com/careers","Wealth AI","70"),
  ("INDmoney","Indian Product","Fintech","High","SDE,ML Eng,Data Sci","14-34 LPA","Gurugram","Yes","0-4 yrs","indmoney.com/careers","Super investing app","78"),
  ("Kuvera","Indian Product","Fintech","Medium","SDE,Backend","12-28 LPA","Bengaluru","Yes","0-3 yrs","kuvera.in/careers","Mutual fund AI","69"),
  ("Fisdom","Indian Product","Fintech","Medium","SDE,Backend","12-28 LPA","Bengaluru","Yes","0-3 yrs","fisdom.com/careers","Wealth platform","68"),
  ("Nuo","Indian Product","Fintech","Low","SDE,Backend","10-22 LPA","Bengaluru","Yes","0-3 yrs","","Crypto alt-invest","50"),
  ("CoinDCX","Indian Product","Crypto","Low","SDE,Backend,ML Eng","12-28 LPA","Mumbai","Yes","0-3 yrs","coindcx.com/careers","Crypto exchange","55"),
  ("WazirX","Indian Product","Crypto","Low","SDE,Backend","10-24 LPA","Mumbai","Yes","0-3 yrs","wazirx.com/careers","Crypto trading","50"),
  ("CoinSwitch","Indian Product","Crypto","Low","SDE,Backend","10-24 LPA","Bengaluru","Yes","0-3 yrs","coinswitch.co/careers","Crypto app","52"),
  ("Mudrex","Indian Product","Crypto","Low","SDE,Backend","10-22 LPA","Bengaluru","Yes","0-3 yrs","mudrex.com","Algo crypto trading","48"),
  ("Perpule (Amazon)","Indian Product","Retail-Tech","Medium","SDE,ML Eng","14-34 LPA","Bengaluru","Yes","0-3 yrs","amazon.jobs","Checkout AI","70"),
  ("Zetwerk","Indian Product","Manufacturing","High","SDE,ML Eng,Data Sci","14-36 LPA","Bengaluru","Yes","0-4 yrs","zetwerk.com/careers","Manufacturing AI","80"),
  ("Delhivery","Indian Product","Logistics","High","SDE,ML Eng,Data Sci","14-36 LPA","Gurugram,Delhi","Yes","0-4 yrs","delhivery.com/careers","Logistics AI","80"),
  ("Rivigo","Indian Product","Logistics","Medium","SDE,ML Eng","12-28 LPA","Gurugram","Yes","0-3 yrs","rivigo.com/careers","Relay trucking AI","71"),
  ("BlackBuck","Indian Product","Logistics","High","SDE,ML Eng,Data Sci","14-34 LPA","Bengaluru","Yes","0-4 yrs","blackbuck.com/careers","Trucking fintech","78"),
  ("Porter","Indian Product","Logistics","High","Backend,SDE,ML Eng","14-32 LPA","Bengaluru","Yes","0-4 yrs","porter.in/careers","Intracity logistics","77"),
  ("Loadshare","Indian Product","Logistics","Medium","SDE,Data Analyst","12-26 LPA","Bengaluru","Yes","0-3 yrs","loadshare.net/careers","Logistics network","69"),
  ("ShipRocket","Indian Product","Logistics","High","SDE,Backend,ML Eng","12-30 LPA","Delhi NCR","Yes","0-4 yrs","shiprocket.in/careers","D2C shipping","77"),
  ("Ecom Express","Indian Product","Logistics","Medium","SDE,Data Analyst","10-24 LPA","Delhi NCR","Yes","0-3 yrs","ecomexpress.in/careers","Last-mile logistics","68"),
  ("XpressBees","Indian Product","Logistics","Medium","SDE,Data Analyst","10-24 LPA","Pune","Yes","0-3 yrs","xpressbees.com/careers","Express logistics","67"),
  ("Shadowfax","Indian Product","Logistics","Medium","SDE,ML Eng","12-26 LPA","Bengaluru","Yes","0-3 yrs","shadowfax.in/careers","Hyperlocal delivery","69"),
  ("WayCool","Indian Product","AgriTech","Medium","SDE,Data Analyst,ML Eng","12-28 LPA","Chennai","Yes","0-3 yrs","waycoolfoods.com/careers","Agri-supply chain","69"),
  ("Ninjacart","Indian Product","AgriTech","High","SDE,ML Eng,Data Sci","14-32 LPA","Bengaluru","Yes","0-4 yrs","ninjacart.com/careers","Agri B2B AI","77"),
  ("DeHaat","Indian Product","AgriTech","Medium","SDE,Data Analyst","12-26 LPA","Patna,Delhi NCR","Yes","0-3 yrs","dehaat.com/careers","Farmer digital","68"),
  ("AgroStar","Indian Product","AgriTech","Medium","SDE,ML Eng","12-26 LPA","Pune","Yes","0-3 yrs","agrostar.in/careers","Agri advisory AI","67"),
  ("Bijak","Indian Product","AgriTech","Medium","SDE,Data Analyst","12-24 LPA","Gurugram","Yes","0-3 yrs","bijak.in/careers","Agri trade platform","65"),
  ("Cropin","Indian Product","AgriAI","High","SDE,ML Eng,Data Sci","14-34 LPA","Bengaluru","Yes","0-4 yrs","cropin.com/careers","Farm AI platform","77"),
  ("SatSure","Indian Product","AgriAI","Medium","SDE,ML Eng,Data Sci","12-28 LPA","Bengaluru","Yes","0-3 yrs","satsure.co/careers","Satellite agri AI","72"),
  ("Intello Labs","Indian Product","AgriAI","Medium","ML Eng,SDE","12-28 LPA","Gurugram","Yes","0-3 yrs","intellolabs.com/careers","Grain quality AI","70"),
  ("Arya.ag","Indian Product","AgriTech","Medium","SDE,Data Analyst","10-24 LPA","Delhi NCR","Yes","0-3 yrs","arya.ag/careers","Post-harvest AI","65"),
  ("Wadhwani AI","NonProfit-AI","Social-AI","High","ML Eng,Data Sci","14-30 LPA","Mumbai","No","2+ yrs","wadhwaniai.org/careers","Social impact AI","75"),
  ("iMerit","AI Company","Data-Label","High","ML Eng,Data Sci,SDE","12-32 LPA","Kolkata,Bengaluru","Yes","0-4 yrs","imerit.net/careers","AI training data","75"),
  ("Saama Technologies","AI Company","LifeSci-AI","High","ML Eng,Data Sci,SDE","16-40 LPA","Bengaluru,Pune","No","2+ yrs","saama.com/company/careers","Clinical AI","78"),
  ("Tiger Analytics","Analytics","Consulting","High","Data Sci,ML Eng,Data Analyst","14-38 LPA","Chennai,Bengaluru","No","1+ yr","tigeranalytics.com/careers","Advanced analytics","79"),
  ("Fractal Analytics","Analytics","Consulting","Very High","Data Sci,ML Eng,Data Analyst","14-40 LPA","Mumbai,Bengaluru","Yes","0-4 yrs","fractal.ai/careers","AI consulting","84"),
  ("Mu Sigma","Analytics","Consulting","High","Data Sci,ML Eng,Data Analyst","12-32 LPA","Bengaluru","Yes","0-4 yrs","mu-sigma.com/careers","Decision science","80"),
  ("Latent View Analytics","Analytics","Consulting","High","Data Sci,ML Eng,BI Analyst","12-34 LPA","Chennai,Bengaluru","Yes","0-4 yrs","latentview.com/careers","Analytics platform","80"),
  ("AbsolutData","Analytics","Consulting","Medium","Data Sci,ML Eng","12-30 LPA","Gurugram","Yes","0-3 yrs","absolutdata.com/careers","Marketing AI","73"),
  ("EXL Service","Analytics","BPO+AI","Very High","Data Sci,ML Eng,SDE","12-34 LPA","Delhi NCR,Noida","Yes","0-4 yrs","exlservice.com/careers","AI-driven BPO","83"),
  ("WNS Global","Analytics","BPO+AI","High","Data Sci,ML Eng,BI Analyst","10-28 LPA","Mumbai,Pune","Yes","0-4 yrs","wns.com/careers","Analytics ops","79"),
  ("Genpact","Analytics","BPO+AI","Very High","Data Sci,ML Eng,SDE","10-28 LPA","Delhi NCR,Hyderabad","Yes","0-4 yrs","genpact.com/careers","AI transformation","82"),
  ("Conduent","Analytics","BPO","Medium","SDE,Data Analyst","10-24 LPA","Hyderabad,Bengaluru","Yes","0-3 yrs","conduent.com/careers","Digital ops","67"),
  ("iGate (Capgemini)","Service","IT","Medium","SDE,Backend","10-26 LPA","Pune","Yes","0-3 yrs","capgemini.com/careers","Capgemini India","70"),
  ("MindTree (LTI)","Service","IT","High","SDE,ML Eng,Data Sci","12-34 LPA","Bengaluru","Yes","0-4 yrs","ltimindtree.com/careers","LTI merger synergy","79"),
  ("Infosys BPM","Service","BPO","High","SDE,Data Sci,BI Analyst","8-22 LPA","Bengaluru,Pune","Yes","0-4 yrs","infosysbpm.com/careers","AI ops","77"),
  ("Wipro BPS","Service","BPO","High","SDE,Data Analyst","8-20 LPA","Chennai,Hyderabad","Yes","0-4 yrs","wipro.com/careers","AI-assisted ops","76"),
  ("Conduent India","Service","BPO","Medium","SDE,Data Analyst","8-20 LPA","Hyderabad","Yes","0-3 yrs","conduent.com/careers","Digital services","65"),
  ("iEnergizer","Service","BPO","Low","Data Analyst,SDE","6-14 LPA","Delhi NCR","Yes","0-3 yrs","ienergizer.com/careers","BPO services","48"),
  ("NIIT Data","EdTech-AI","EdTech","Low","SDE,Data Analyst","6-14 LPA","Delhi NCR","Yes","0-2 yrs","niit.com","Skill training","45"),
  ("TestBook","Indian Product","EdTech","High","SDE,Backend,ML Eng","12-30 LPA","Delhi NCR","Yes","0-4 yrs","testbook.com/careers","Govt exam AI","78"),
  ("Doubtnut","Indian Product","EdTech","Medium","SDE,ML Eng","12-26 LPA","Delhi NCR","Yes","0-3 yrs","doubtnut.com/careers","Math AI","69"),
  ("Extramarks","Indian Product","EdTech","Medium","SDE,ML Eng,Data Analyst","10-24 LPA","Delhi NCR","Yes","0-3 yrs","extramarks.com/careers","K-12 AI","66"),
  ("iQuanta","Indian Product","EdTech","Low","SDE,Backend","8-18 LPA","Delhi NCR","Yes","0-3 yrs","iquanta.in/careers","CAT prep","50"),
  ("Careers360","Indian Product","EdTech","Medium","SDE,Data Analyst,ML Eng","10-24 LPA","Delhi NCR","Yes","0-3 yrs","careers360.com/careers","College search AI","66"),
  ("Collegedunia","Indian Product","EdTech","Medium","SDE,Backend","10-22 LPA","Delhi NCR","Yes","0-3 yrs","collegedunia.com/careers","College discovery","63"),
  ("Shiksha (InfoEdge)","Indian Product","EdTech","Medium","SDE,Data Analyst","10-22 LPA","Noida","Yes","0-3 yrs","infoedge.in/careers","Education portal","63"),
  ("MyCaptain","Indian Product","EdTech","Low","SDE,Backend","8-16 LPA","Bengaluru","Yes","0-2 yrs","mycaptain.in/careers","Career guidance","48"),
  ("Olern","Indian Product","EdTech","Low","SDE,Backend","8-14 LPA","Delhi NCR","Yes","0-2 yrs","olern.com","Corporate learning","44"),
]

# Write rows
cat_colors = {
    "Product": (LIGHT_BLUE, WHITE), "GCC": (TEAL, WHITE),
    "Service": (MID_GRAY, DARK_GRAY), "Indian Product": ("1A237E", WHITE),
    "AI Startup": (PURPLE, WHITE), "Unicorn": (ORANGE, WHITE),
    "Fintech": ("880E4F", WHITE), "SaaS": (GREEN, WHITE),
    "AI Company": (PURPLE, WHITE), "Series A": (LIGHT_GREEN, DARK_GRAY),
    "Series B": (LIGHT_GREEN, DARK_GRAY), "Series C": (LIGHT_TEAL, DARK_GRAY),
    "Analytics": (ACCENT_GOLD, DARK_GRAY), "EdTech": (LIGHT_ORANGE, DARK_GRAY),
    "Neobank": ("4A148C", WHITE), "Swedish Product": (LIGHT_BLUE, DARK_GRAY),
    "NonProfit-AI": (LIGHT_TEAL, DARK_GRAY),
}

for i, c in enumerate(companies, 1):
    row_num = i + 2
    row_data = [i] + list(c)
    ws1.row_dimensions[row_num].height = 22
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for col_j, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_num, column=col_j, value=val)
        cell.border = thin_border()
        cell.alignment = left()
        cell.font = font(False, 9)
        # Category color badge on column 3
        if col_j == 3:
            cat = c[1]
            cf, ct = cat_colors.get(cat, (bg, DARK_GRAY))
            cell.fill = fill(cf)
            cell.font = font(True, 9, ct)
        elif col_j == 5:
            prob = c[3]
            pf = {"Very High": GREEN, "High": LIGHT_BLUE, "Medium": ACCENT_GOLD,
                  "Low": RED}.get(prob, bg)
            cell.fill = fill(pf)
            cell.font = font(True, 9, WHITE if prob in ("Very High","High","Low") else DARK_GRAY)
        elif col_j == 9:
            cell.fill = fill(LIGHT_GREEN if val == "Yes" else LIGHT_RED)
            cell.font = font(True, 9, GREEN if val == "Yes" else RED)
        else:
            cell.fill = fill(bg)

# Auto-filter
ws1.auto_filter.ref = f"A2:{get_column_letter(14)}{len(companies)+2}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INTERVIEW QUESTIONS
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📝 Interview Questions")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value = "🧠  INTERVIEW QUESTION BANK — SDE | Backend | ML | Data Analyst | Data Scientist | MLOps"
t2.fill = fill(DARK_NAVY)
t2.font = Font(bold=True, size=13, color=ACCENT_GOLD, name="Calibri")
t2.alignment = center()
ws2.row_dimensions[1].height = 34

q_cols = ["#","Domain","Topic","Difficulty","Question","Company Tags"]
q_fills = [DARK_NAVY,DARK_NAVY,DARK_NAVY,DARK_NAVY,DARK_NAVY,DARK_NAVY]
q_fonts = [font(True,10,WHITE)]*6
header_row(ws2, 2, q_cols, q_fills, q_fonts)
for i,w in enumerate([4,18,18,12,80,28],1):
    set_col_width(ws2, get_column_letter(i), w)

questions = [
    # SDE / DSA
    ("SDE/DSA","Arrays","Easy","Two Sum — find two indices in array that add to target","Google,Amazon,Microsoft"),
    ("SDE/DSA","Arrays","Medium","Best Time to Buy and Sell Stock (Kadane's variant)","Adobe,Flipkart,Uber"),
    ("SDE/DSA","Arrays","Medium","Subarray with given sum (sliding window)","Amazon,Walmart"),
    ("SDE/DSA","Arrays","Hard","Trapping Rainwater — two-pointer approach","Google,Meta,Razorpay"),
    ("SDE/DSA","Arrays","Hard","Merge K sorted arrays using min-heap","Microsoft,Flipkart"),
    ("SDE/DSA","Strings","Easy","Check if two strings are anagrams","Zoho,Infosys"),
    ("SDE/DSA","Strings","Medium","Longest Palindromic Substring (DP or Manacher's)","Amazon,Google"),
    ("SDE/DSA","Strings","Medium","Group anagrams together from list of strings","Atlassian,Salesforce"),
    ("SDE/DSA","Strings","Hard","Minimum window substring","Meta,Uber"),
    ("SDE/DSA","Strings","Hard","Regular expression matching (DP)","Google,Microsoft"),
    ("SDE/DSA","Linked List","Easy","Reverse a linked list (iterative & recursive)","All companies"),
    ("SDE/DSA","Linked List","Medium","Detect cycle in linked list (Floyd's algorithm)","Amazon,Wipro"),
    ("SDE/DSA","Linked List","Medium","Merge two sorted linked lists","TCS,Infosys,Wipro"),
    ("SDE/DSA","Linked List","Hard","Copy list with random pointer","Microsoft,Google"),
    ("SDE/DSA","Linked List","Hard","LRU Cache implementation","Amazon,Flipkart,Groww"),
    ("SDE/DSA","Trees","Easy","Inorder, Preorder, Postorder traversal","All companies"),
    ("SDE/DSA","Trees","Medium","Lowest Common Ancestor of BST","Google,Amazon"),
    ("SDE/DSA","Trees","Medium","Zigzag level order traversal","Walmart,Target"),
    ("SDE/DSA","Trees","Hard","Serialize and deserialize binary tree","Meta,Uber,Swiggy"),
    ("SDE/DSA","Trees","Hard","Maximum path sum in binary tree","Google,Microsoft"),
    ("SDE/DSA","Graphs","Medium","Number of islands (BFS/DFS)","Amazon,Flipkart"),
    ("SDE/DSA","Graphs","Medium","Shortest path in unweighted graph (BFS)","Dunzo,Porter"),
    ("SDE/DSA","Graphs","Medium","Detect cycle in directed graph","Goldman Sachs"),
    ("SDE/DSA","Graphs","Hard","Dijkstra's algorithm implementation","Uber,Ola,Google Maps"),
    ("SDE/DSA","Graphs","Hard","Alien Dictionary — topological sort","Airbnb,Codeforces"),
    ("SDE/DSA","Dynamic Programming","Medium","0/1 Knapsack","Accenture,TCS"),
    ("SDE/DSA","Dynamic Programming","Medium","Longest Increasing Subsequence","Amazon,Morgan Stanley"),
    ("SDE/DSA","Dynamic Programming","Hard","Edit Distance (Levenshtein)","Microsoft,Google"),
    ("SDE/DSA","Dynamic Programming","Hard","Burst Balloons","Meta (advanced)"),
    ("SDE/DSA","Dynamic Programming","Hard","Palindrome Partitioning","Google,Samsung"),
    ("SDE/DSA","Heaps/Priority Q","Medium","Kth largest element","Amazon,Flipkart,Razorpay"),
    ("SDE/DSA","Heaps/Priority Q","Hard","Find median from data stream","Google,JP Morgan"),
    ("SDE/DSA","Trie","Medium","Implement Trie (insert, search, startswith)","Google,Microsoft,Postman"),
    ("SDE/DSA","Backtracking","Medium","Generate all permutations","All FAANG"),
    ("SDE/DSA","Backtracking","Hard","N-Queens problem","Google,Meta"),
    ("SDE/DSA","Bit Manipulation","Medium","Count set bits in an integer","Texas Instruments"),
    ("SDE/DSA","Sliding Window","Medium","Longest substring without repeating characters","All companies"),
    ("SDE/DSA","Two Pointer","Medium","Three Sum (find triplets summing to zero)","Adobe,Salesforce"),
    ("SDE/DSA","Binary Search","Medium","Search in rotated sorted array","Google,Amazon"),
    ("SDE/DSA","Binary Search","Hard","Median of two sorted arrays","Google (hard)"),
    # Backend / Python
    ("Backend","Python","Easy","Explain GIL in Python and when it matters","Zerodha,Razorpay"),
    ("Backend","Python","Medium","Difference between multiprocessing vs multithreading in Python","Groww,CRED"),
    ("Backend","Python","Medium","How does Python's async/await work? What is event loop?","Postman,Stripe"),
    ("Backend","Python","Medium","Explain Python decorators with a real-world example","Swiggy,Zomato"),
    ("Backend","Python","Hard","Design a rate limiter in Python using Redis","Razorpay,Stripe,Juspay"),
    ("Backend","Django/FastAPI","Easy","What are Django signals and when to use them?","TCS,Wipro"),
    ("Backend","Django/FastAPI","Medium","Explain async views in Django 4.x and ASGI","Freshworks,Zoho"),
    ("Backend","Django/FastAPI","Medium","How does FastAPI dependency injection work?","Zerodha,Chargebee"),
    ("Backend","Django/FastAPI","Medium","Explain FastAPI's Pydantic models and validation","Hasura,Appsmith"),
    ("Backend","Django/FastAPI","Hard","Design a multi-tenant SaaS architecture with FastAPI","Postman,Atlassian"),
    ("Backend","APIs","Medium","REST vs GraphQL — when to use which?","Hasura,Postman,Github"),
    ("Backend","APIs","Medium","Explain idempotency in REST APIs with examples","Stripe,Razorpay"),
    ("Backend","APIs","Hard","Design a webhook delivery system with retry logic","Razorpay,Cashfree"),
    ("Backend","APIs","Hard","API versioning strategies — URI, header, query param","Salesforce,Adobe"),
    ("Backend","PostgreSQL","Easy","Explain ACID properties with examples","All companies"),
    ("Backend","PostgreSQL","Medium","B-tree vs Hash indexes in PostgreSQL","Zerodha,JP Morgan"),
    ("Backend","PostgreSQL","Medium","Explain window functions with example","Goldman Sachs,Flipkart"),
    ("Backend","PostgreSQL","Medium","How would you optimize a slow query (EXPLAIN ANALYZE)?","Groww,PhonePe"),
    ("Backend","PostgreSQL","Hard","Design a schema for a multi-currency payment system","Razorpay,PayU"),
    ("Backend","Redis","Medium","When to use Redis sorted sets vs lists?","Zomato,Swiggy"),
    ("Backend","Redis","Medium","Implement distributed locks using Redis (SETNX)","Flipkart,Amazon"),
    ("Backend","Redis","Hard","Redis cluster vs Redis Sentinel — trade-offs","CRED,Chargebee"),
    ("Backend","Kafka","Medium","Explain Kafka partitions, offsets and consumer groups","Swiggy,Meesho"),
    ("Backend","Kafka","Medium","At-least-once vs exactly-once semantics in Kafka","PhonePe,Delhivery"),
    ("Backend","Kafka","Hard","Design an event-sourcing system with Kafka","Flipkart,Paytm"),
    ("Backend","System Design","Hard","Design URL Shortener (like bit.ly)","All companies"),
    ("Backend","System Design","Hard","Design WhatsApp messaging system","Meta,Microsoft"),
    ("Backend","System Design","Hard","Design Uber/Ola ride-matching system","Uber,Ola,Rapido"),
    ("Backend","System Design","Hard","Design Netflix content delivery system","Netflix,Hotstar"),
    ("Backend","System Design","Hard","Design a distributed cache (like Memcached)","Amazon,Google"),
    # ML / Data Science
    ("ML/DS","ML Fundamentals","Easy","Explain bias-variance tradeoff with an example","All data roles"),
    ("ML/DS","ML Fundamentals","Medium","When would you use L1 vs L2 regularization?","Fractal,EXL,Tiger Analytics"),
    ("ML/DS","ML Fundamentals","Medium","Explain gradient boosting vs random forest","Walmart,Target,Flipkart"),
    ("ML/DS","ML Fundamentals","Hard","Derive the backpropagation algorithm from scratch","Google,DeepMind,NVIDIA"),
    ("ML/DS","ML Fundamentals","Hard","How do you handle class imbalance in production?","PhonePe,Razorpay"),
    ("ML/DS","Transformers/LLMs","Medium","Explain self-attention mechanism in transformers","Sarvam AI,Krutrim"),
    ("ML/DS","Transformers/LLMs","Medium","What is LoRA fine-tuning and when to use it?","Hugging Face,Cohere"),
    ("ML/DS","Transformers/LLMs","Hard","Design a RAG pipeline for enterprise search","Atlassian,Salesforce"),
    ("ML/DS","Transformers/LLMs","Hard","How would you evaluate an LLM in production?","Galileo,Arize AI"),
    ("ML/DS","Transformers/LLMs","Hard","Explain RLHF and its challenges","OpenAI,Anthropic"),
    ("ML/DS","Statistics","Medium","Explain p-value and its common misinterpretations","JP Morgan,Goldman Sachs"),
    ("ML/DS","Statistics","Medium","Central Limit Theorem and its applications in ML","Fractal,Mu Sigma"),
    ("ML/DS","Statistics","Medium","A/B testing — how do you determine sample size?","Flipkart,Amazon,Booking"),
    ("ML/DS","Statistics","Hard","Explain Bayesian inference vs frequentist approach","Google,Meta"),
    ("ML/DS","Statistics","Hard","How to detect and handle distribution shift in prod?","PhonePe,Swiggy"),
    ("ML/DS","Python/ML","Medium","Implement k-means clustering from scratch (NumPy)","Fractal,Tiger Analytics"),
    ("ML/DS","Python/ML","Medium","Explain difference between .fit(), .transform(), .predict()","TCS (digital),Wipro"),
    ("ML/DS","Python/ML","Hard","Build a mini neural net using only NumPy","Google,NVIDIA"),
    ("ML/DS","SQL for DS","Medium","Find the median salary per department","All companies"),
    ("ML/DS","SQL for DS","Medium","Write SQL for rolling 7-day average of sales","Walmart,Target"),
    ("ML/DS","SQL for DS","Hard","Find users who made purchases on 3 consecutive days","Amazon,Flipkart"),
    ("ML/DS","SQL for DS","Hard","Design a feature store schema for ML features","Databricks,Snowflake"),
    # Data Analyst
    ("Data Analyst","SQL","Easy","Difference between WHERE and HAVING clause","All companies"),
    ("Data Analyst","SQL","Medium","Explain and use RANK(), DENSE_RANK(), ROW_NUMBER()","Goldman Sachs,JP Morgan"),
    ("Data Analyst","SQL","Medium","Self-join to find employees earning more than their managers","Microsoft,Wipro"),
    ("Data Analyst","SQL","Hard","Recursive CTE for org hierarchy","Oracle,SAP"),
    ("Data Analyst","Excel/Sheets","Medium","VLOOKUP vs INDEX-MATCH — when to use each?","Accenture,Cognizant"),
    ("Data Analyst","Power BI","Medium","Explain DAX CALCULATE function with example","Wipro,Infosys"),
    ("Data Analyst","Power BI","Hard","Design a star schema for a retail analytics dashboard","WNS,EXL"),
    ("Data Analyst","Statistics","Medium","Explain correlation vs causation with business example","Fractal,Mu Sigma"),
    ("Data Analyst","Case Study","Hard","Drop in conversion rate — how do you diagnose it?","Flipkart,Meesho"),
    ("Data Analyst","Case Study","Hard","Recommend a KPI framework for an e-commerce company","Amazon,Walmart"),
    # MLOps
    ("MLOps","MLOps","Medium","What is a feature store and why use one?","Databricks,Feast"),
    ("MLOps","MLOps","Medium","Explain ML model versioning strategies","Weights & Biases,MLflow"),
    ("MLOps","MLOps","Hard","Design a CI/CD pipeline for ML models","Netflix,Uber,Swiggy"),
    ("MLOps","MLOps","Hard","How do you monitor model drift in production?","Arize AI,Fiddler AI"),
    ("MLOps","Cloud/Infra","Medium","Explain Kubernetes concepts relevant to ML workloads","Google,Microsoft"),
    ("MLOps","Cloud/Infra","Hard","Design a distributed training pipeline for LLMs","NVIDIA,Google,Cohere"),
    ("MLOps","Docker/K8s","Medium","What is a sidecar container pattern?","Atlassian,Harman"),
    ("MLOps","Docker/K8s","Hard","Explain Kubernetes HPA and KEDA for ML serving","Scale AI,Together AI"),
]

diff_colors = {"Easy": GREEN, "Medium": ACCENT_GOLD, "Hard": RED}
domain_colors = {
    "SDE/DSA": MID_BLUE, "Backend": TEAL, "ML/DS": PURPLE,
    "Data Analyst": ORANGE, "MLOps": "37474F"
}
for i, q in enumerate(questions, 1):
    row_num = i + 2
    ws2.row_dimensions[row_num].height = 28
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    row_data = [i, q[0], q[1], q[2], q[3], q[4]]
    for col_j, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_num, column=col_j, value=val)
        cell.border = thin_border()
        cell.alignment = left()
        cell.font = font(False, 9)
        if col_j == 2:
            dc = domain_colors.get(q[0], MID_BLUE)
            cell.fill = fill(dc)
            cell.font = font(True, 9, WHITE)
        elif col_j == 4:
            cell.fill = fill(diff_colors.get(q[2], bg))
            cell.font = font(True, 9, WHITE)
        else:
            cell.fill = fill(bg)

ws2.auto_filter.ref = f"A2:F{len(questions)+2}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SALARY DATABASE
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("💰 Salary Database")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
t3 = ws3["A1"]
t3.value = "💰  SALARY INTELLIGENCE DATABASE — India Tech Hiring 2026  |  All figures in LPA (Lakhs Per Annum)"
t3.fill = fill(DARK_NAVY)
t3.font = Font(bold=True, size=13, color=ACCENT_GOLD, name="Calibri")
t3.alignment = center()
ws3.row_dimensions[1].height = 34

sal_cols = ["Role","Level","Company Type","Min (LPA)","Median (LPA)","Avg (LPA)","Top 10% (LPA)","Notes"]
for i, w in enumerate([22,18,20,12,14,14,14,30], 1):
    set_col_width(ws3, get_column_letter(i), w)
header_row(ws3, 2, sal_cols,
           [MID_BLUE]*8, [font(True,10,WHITE)]*8)

salary_data = [
    ("SDE","Fresher (0-1 yr)","Service (TCS/Infy/Wipro)","3.5","4.5","4.2","6.5","Ninja/Digital roles at 6-8 LPA"),
    ("SDE","Fresher (0-1 yr)","GCC (Walmart/JP Morgan)","8","14","12","20","Grade-based intake"),
    ("SDE","Fresher (0-1 yr)","Indian Product (Zerodha/Razorpay)","10","18","16","28","Stock options critical"),
    ("SDE","Fresher (0-1 yr)","Global FAANG","25","35","32","50","Rare direct fresher; mainly internship converts"),
    ("SDE","SDE-1 (1-3 yr)","Service","5","8","7.5","12","Variable skills-based bands"),
    ("SDE","SDE-1 (1-3 yr)","GCC","12","20","18","30","Base + bonus"),
    ("SDE","SDE-1 (1-3 yr)","Indian Product","14","24","22","35","ESOP + variable"),
    ("SDE","SDE-1 (1-3 yr)","Global FAANG","28","40","38","60","RSU vesting significant"),
    ("SDE","SDE-2 (3-6 yr)","Service","8","15","14","22",""),
    ("SDE","SDE-2 (3-6 yr)","GCC","18","32","30","48",""),
    ("SDE","SDE-2 (3-6 yr)","Indian Product","22","38","36","55",""),
    ("SDE","SDE-2 (3-6 yr)","Global FAANG","40","65","60","90",""),
    ("Backend Developer (Python)","Fresher","Service","3.5","4.5","4.2","6","Django/Flask knowledge"),
    ("Backend Developer (Python)","Fresher","Indian Startup","10","16","14","24","FastAPI hot in 2026"),
    ("Backend Developer (Python)","Fresher","GCC","8","14","12","20",""),
    ("Backend Developer (Python)","Mid (2-4 yr)","Indian Product","16","28","26","40","Microservices exp key"),
    ("Backend Developer (Python)","Mid (2-4 yr)","GCC","18","32","30","45",""),
    ("Backend Developer (Python)","Mid (2-4 yr)","Global Product","25","42","40","60",""),
    ("Backend Developer (Python)","Senior (5+ yr)","Indian Product","32","52","48","75","System design critical"),
    ("Backend Developer (Python)","Senior (5+ yr)","GCC","38","60","55","85",""),
    ("ML Engineer","Fresher","Service/Analytics","5","8","7","12","MSc/BTech AI"),
    ("ML Engineer","Fresher","AI Startup","14","22","20","32","LLM/GenAI premium"),
    ("ML Engineer","Fresher","Indian Product","16","26","24","38",""),
    ("ML Engineer","Fresher","Global Product","26","40","38","58",""),
    ("ML Engineer","Mid (2-4 yr)","AI Startup","22","38","36","55","RAG/Fine-tune exp"),
    ("ML Engineer","Mid (2-4 yr)","Indian Product","22","36","34","52",""),
    ("ML Engineer","Mid (2-4 yr)","GCC","24","40","38","58",""),
    ("ML Engineer","Mid (2-4 yr)","Global FAANG","40","65","60","90",""),
    ("ML Engineer","Senior (5+ yr)","Indian Product","38","60","55","80",""),
    ("ML Engineer","Senior (5+ yr)","Global FAANG","55","90","85","130",""),
    ("Data Analyst","Fresher","Service","3.5","5","4.8","7","SQL + Excel + Power BI"),
    ("Data Analyst","Fresher","GCC","7","12","10","18","Analytics infra roles"),
    ("Data Analyst","Fresher","Indian Product","8","14","12","20",""),
    ("Data Analyst","Mid (2-4 yr)","Service","6","10","9","14",""),
    ("Data Analyst","Mid (2-4 yr)","GCC","12","20","18","28",""),
    ("Data Analyst","Mid (2-4 yr)","Indian Product","14","24","22","35",""),
    ("Data Analyst","Senior (5+ yr)","Indian Product","22","36","34","50",""),
    ("Data Analyst","Senior (5+ yr)","GCC","26","42","40","58",""),
    ("Data Scientist","Fresher","Analytics Firm","6","10","9","14","Statistics + Python"),
    ("Data Scientist","Fresher","Indian Product","12","20","18","28","Kaggle/research exp"),
    ("Data Scientist","Fresher","GCC","12","20","18","28",""),
    ("Data Scientist","Mid (2-4 yr)","Indian Product","20","35","32","48",""),
    ("Data Scientist","Mid (2-4 yr)","GCC","22","38","36","52",""),
    ("Data Scientist","Mid (2-4 yr)","Global Product","32","52","48","72",""),
    ("Data Scientist","Senior (5+ yr)","Indian Product","35","55","50","75",""),
    ("Data Scientist","Senior (5+ yr)","Global FAANG","52","85","80","120",""),
    ("MLOps Engineer","Mid (2-4 yr)","Indian Product","18","30","28","42","Docker/K8s/MLflow"),
    ("MLOps Engineer","Mid (2-4 yr)","GCC","20","34","32","48",""),
    ("MLOps Engineer","Mid (2-4 yr)","Global Product","28","48","45","68",""),
    ("MLOps Engineer","Senior (5+ yr)","Indian Product","32","52","48","70",""),
    ("MLOps Engineer","Senior (5+ yr)","Global Product","48","78","72","110",""),
    ("BI Analyst","Fresher","Service","3.5","5","4.8","7","Power BI + SQL + Tableau"),
    ("BI Analyst","Fresher","GCC","7","12","10","16",""),
    ("BI Analyst","Mid (2-4 yr)","GCC","12","20","18","28",""),
    ("BI Analyst","Mid (2-4 yr)","Indian Product","10","18","16","26",""),
    ("BI Analyst","Senior (5+ yr)","GCC","20","34","30","46",""),
    ("Internship","Summer Intern","Service","10k-20k/mo","15k/mo","14k/mo","25k/mo","Monthly stipend"),
    ("Internship","Summer Intern","Indian Product","30k-60k/mo","45k/mo","42k/mo","75k/mo","PPO likely at top cos"),
    ("Internship","Summer Intern","FAANG","80k-1.2L/mo","1L/mo","95k/mo","1.4L/mo","PPO = direct offer"),
    ("Internship","Summer Intern","GCC","30k-60k/mo","45k/mo","40k/mo","70k/mo",""),
]

level_colors = {
    "Fresher (0-1 yr)": LIGHT_GREEN, "Fresher": LIGHT_GREEN,
    "SDE-1 (1-3 yr)": LIGHT_BLUE, "Mid (2-4 yr)": LIGHT_BLUE,
    "SDE-2 (3-6 yr)": LIGHT_TEAL, "Senior (5+ yr)": LIGHT_ORANGE,
    "Summer Intern": LIGHT_PURPLE,
}
for i, s in enumerate(salary_data, 1):
    row_num = i + 2
    ws3.row_dimensions[row_num].height = 22
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for col_j, val in enumerate([i] + list(s), 1):
        if col_j == 1:
            continue  # skip index col
        cell = ws3.cell(row=row_num, column=col_j-1, value=val)
        cell.border = thin_border()
        cell.alignment = left()
        cell.font = font(False, 9)
        lev_bg = level_colors.get(s[1], bg) if col_j == 3 else bg
        cell.fill = fill(lev_bg)

ws3.auto_filter.ref = f"A2:H{len(salary_data)+2}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — APPLICATION TRACKER
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📋 Application Tracker")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A3"

ws4.merge_cells("A1:L1")
t4 = ws4["A1"]
t4.value = "📋  APPLICATION TRACKER — Aug–Dec 2026  |  Track Every Application to Maximize Offer Rate"
t4.fill = fill(DARK_NAVY)
t4.font = Font(bold=True, size=13, color=ACCENT_GOLD, name="Calibri")
t4.alignment = center()
ws4.row_dimensions[1].height = 34

tr_cols = ["#","Company","Role Applied","Date Applied","OA?","Technical 1","Technical 2","HR Round",
           "Final Status","Offer (LPA)","Follow-up Date","Notes / Referral"]
tr_fills = [DARK_NAVY]*12
tr_fonts = [font(True,10,WHITE)]*12
header_row(ws4, 2, tr_cols, tr_fills, tr_fonts)
for i, w in enumerate([4,24,22,14,10,14,14,12,14,12,14,30], 1):
    set_col_width(ws4, get_column_letter(i), w)

# Pre-fill 50 rows with top companies
top50 = [c[0] for c in companies[:50]]
statuses = ["Not Applied","Applied","OA Done","Tech 1 Done","Tech 2 Done","HR Done","Offer","Rejected","Waitlisted"]
for i, comp in enumerate(top50, 1):
    row_num = i + 2
    ws4.row_dimensions[row_num].height = 22
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    row_vals = [i, comp, "", "", "No", "Pending", "Pending", "Pending", "Not Applied", "", "", ""]
    for col_j, val in enumerate(row_vals, 1):
        cell = ws4.cell(row=row_num, column=col_j, value=val)
        cell.border = thin_border()
        cell.alignment = left()
        cell.font = font(False, 9)
        cell.fill = fill(bg)


for i in range(51, 201):
    row_num = i + 2
    ws4.row_dimensions[row_num].height = 20
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for col_j in range(1, 13):
        cell = ws4.cell(row=row_num, column=col_j, value="")
        cell.border = thin_border()
        cell.fill = fill(bg)

ws4.auto_filter.ref = f"A2:L202"

ws5 = wb.create_sheet("🏆 Hiring Priority Score")
ws5.sheet_view.showGridLines = False
ws5.freeze_panes = "A3"

ws5.merge_cells("A1:I1")
t5 = ws5["A1"]
t5.value = "🏆  HIRING PRIORITY SCORE — Ranked by Composite Score (Funding + Growth + AI + GCC + Prob)"
t5.fill = fill(DARK_NAVY)
t5.font = Font(bold=True, size=13, color=ACCENT_GOLD, name="Calibri")
t5.alignment = center()
ws5.row_dimensions[1].height = 34

sc_cols = ["Rank","Company","Category","Hiring Prob Score (25)","Growth Score (20)","Funding/Scale (20)",
           "AI Hiring Score (20)","GCC/India Signal (15)","TOTAL SCORE (100)"]
sc_fills = [MID_BLUE]*9
sc_fonts = [font(True,10,WHITE)]*9
header_row(ws5, 2, sc_cols, sc_fills, sc_fonts)
for i, w in enumerate([6,28,20,18,14,14,16,18,18], 1):
    set_col_width(ws5, get_column_letter(i), w)


scored = []
prob_map = {"Very High": 25, "High": 20, "Medium": 13, "Low": 6}
for c in companies:
    name, cat, subcat, prob, roles, sal, locs, fresh, exp, url, signal, ps = c
    hp = prob_map.get(prob, 13)
    
    total_str = int(ps)
    gs = min(20, max(0, int((total_str - hp) * 20 / 75)))
    fs = min(20, max(0, int(total_str * 0.20)))
    ai_kw = any(k in (roles + signal + cat).lower() for k in ["ml","ai","llm","data sci","mlops","genai"])
    ai_s = 20 if ai_kw else 12
    gcc_kw = any(k in (cat + signal).lower() for k in ["gcc","india expansion","centre","center"])
    gcc_s = 15 if gcc_kw else 8
    total = min(100, hp + gs + fs + ai_s + gcc_s)
    scored.append((name, cat, hp, gs, fs, ai_s, gcc_s, total))

scored.sort(key=lambda x: -x[7])

for rank, s in enumerate(scored, 1):
    row_num = rank + 2
    ws5.row_dimensions[row_num].height = 22
    bg = LIGHT_GRAY if rank % 2 == 0 else WHITE
    row_vals = [rank, s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7]]
    for col_j, val in enumerate(row_vals, 1):
        cell = ws5.cell(row=row_num, column=col_j, value=val)
        cell.border = thin_border()
        cell.alignment = center() if col_j > 3 else left()
        cell.font = font(False, 9)
        if col_j == 9:
            score = val
            if score >= 90:
                cell.fill = fill(GREEN); cell.font = font(True, 10, WHITE)
            elif score >= 80:
                cell.fill = fill(LIGHT_BLUE); cell.font = font(True, 10, WHITE)
            elif score >= 70:
                cell.fill = fill(ACCENT_GOLD); cell.font = font(True, 9, DARK_GRAY)
            elif score >= 60:
                cell.fill = fill(LIGHT_ORANGE); cell.font = font(False, 9)
            else:
                cell.fill = fill(LIGHT_RED)
        elif rank <= 10 and col_j == 1:
            cell.fill = fill(ACCENT_GOLD)
            cell.font = font(True, 10, DARK_GRAY)
        else:
            cell.fill = fill(bg)

ws5.auto_filter.ref = f"A2:{get_column_letter(9)}{len(scored)+2}"


out = "output/India_Tech_Hiring_Intelligence_2026.xlsx"
wb.save(out)
print(f"Excel saved: {out}")
print(f"Total companies: {len(companies)}")
print(f"Total questions: {len(questions)}")
print(f"Total salary rows: {len(salary_data)}")